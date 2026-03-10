#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""检查预期 Excel 文件中的 ILF 类型数据"""
import openpyxl

wb = openpyxl.load_workbook('/Users/linziwang/PycharmProjects/wordToWord/test/fpa/FPA预估2025版V3.xlsx', data_only=True)
print("工作表:", wb.sheetnames)

ws = wb['2. 规模估算']

print("\n查找 ILF 类型的行:")
for row_idx in range(10, 80):
    category = ws.cell(row=row_idx, column=7).value
    if category == 'ILF':
        row_data = [ws.cell(row=row_idx, column=col).value for col in range(1, 13)]
        print(f"行{row_idx}: {row_data}")
