#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import os

# 查找文件
target_dir = 'test/fpa/'
files = [f for f in os.listdir(target_dir) if f.endswith('.xlsx') and '评审版本' in f]

if not files:
    print("未找到期望的 Excel 文件")
    print(f"目录内容：{os.listdir(target_dir)}")
    exit(1)

excel_file = target_dir + files[0]
print(f"读取文件：{excel_file}")
print()

# 读取 Excel
wb = openpyxl.load_workbook(excel_file)
print(f"工作表列表：{wb.sheetnames}")
print()

# 检查每个工作表
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"工作表：{sheet_name}")
    
    # 获取所有非空行
    rows = list(ws.iter_rows(values_only=True))
    non_empty_rows = [row for row in rows if any(cell is not None for cell in row)]
    
    print(f"行数：{len(non_empty_rows)}")
    print(f"列数：{len(max(non_empty_rows, key=len)) if non_empty_rows else 0}")
    
    if non_empty_rows:
        print(f"前 3 行数据:")
        for i, row in enumerate(non_empty_rows[:3], 1):
            # 过滤掉 None 值
            clean_row = [str(cell) if cell is not None else '' for cell in row]
            print(f"  行{i}: {clean_row[:5]}...")  # 只显示前 5 列
    
    print("-" * 80)
    print()
