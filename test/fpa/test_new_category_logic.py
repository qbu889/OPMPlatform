#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys
import os

# 添加项目根目录到路径
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)
os.chdir(project_root)  # 切换到项目根目录

from routes.fpa_generator_routes import parse_requirement_document, generate_fpa_excel

# 读取测试文档
with open('test/fpa/事件处理类工单流程应用场景.md', 'r', encoding='utf-8') as f:
    content = f.read()

# 解析并生成
print("正在解析需求文档...")
fps = parse_requirement_document(content)
print(f"解析成功，共 {len(fps)} 个功能点")

# 统计类别分布
categories = {}
for fp in fps:
    cat = fp.get('类别', '未知')
    categories[cat] = categories.get(cat, 0) + 1

print(f"\n类别分布:")
for cat, count in sorted(categories.items()):
    print(f"  {cat}: {count} 条")

# 生成 Excel
output_path = 'tmp/test_new_fpa.xlsx'
print(f"\n正在生成 {output_path}...")
generate_fpa_excel(fps, output_path)
print(f"生成成功！")

# 验证生成的 Excel
from openpyxl import load_workbook
wb = load_workbook(output_path)
print(f"\n工作表列表：{wb.sheetnames}")

# 检查"2. 规模估算"表的前几行数据
ws = wb['2. 规模估算']
print(f"\n2. 规模估算 - 前 5 行数据:")
for i in range(9, min(14, ws.max_row + 1)):
    row_data = [ws.cell(row=i, column=j).value for j in range(1, 8)]
    print(f"  行{i}: {row_data}")

print("\n完成！")
