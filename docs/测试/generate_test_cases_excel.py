import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# 重新设计的测试用例数据 - 包含所有要求字段
test_cases = [
    # ========== 模块一：跨专业关联触发逻辑 ==========
    ["TC-001", "跨专业关联触发", "关联追单触发条件验证", "集客 + 家宽跨专业关联追单成功，时限更新",
     "1. 系统已配置集客和家宽专业规则\n2. 关联追单功能已启用",
     "1. 登录系统，创建主事件（集客专业）\n2. 设置主事件故障处理时限=60 分钟\n3. 创建子事件（家宽专业）\n4. 设置子事件故障处理时限=120 分钟\n5. 触发关联追单操作",
     "主事件：专业=集客，时限=60 分钟\n子事件：专业=家宽，时限=120 分钟",
     "1. 系统判断主单专业 (集客) != 子单专业 (家宽)\n2. 判断主单处理时限 (60 分钟) < 子单处理时限 (120 分钟)\n3. 按子单处理时限 + 主单派发时间重新计算\n4. 更新工单及任务处理时限（故障受理/处理时限、T1 受理/处理时限、T2 受理/处理时限）\n5. 任务描述更新为最新一条\n6. 调度作战室留痕成功\n7. IVR 通知当前处理人/组", "P0", "是", "核心正向场景"],
    
    ["TC-002", "跨专业关联触发", "关联追单触发条件验证", "家宽 + 集客跨专业关联追单成功，时限更新",
     "1. 系统已配置集客和家宽专业规则\n2. 关联追单功能已启用",
     "1. 登录系统，创建主事件（家宽专业）\n2. 设置主事件故障处理时限=90 分钟\n3. 创建子事件（集客专业）\n4. 设置子事件故障处理时限=60 分钟\n5. 触发关联追单操作",
     "主事件：专业=家宽，时限=90 分钟\n子事件：专业=集客，时限=60 分钟",
     "1. 系统判断主单专业 (家宽) != 子单专业 (集客)\n2. 判断主单处理时限 (90 分钟) < 子单处理时限 (60 分钟)=False\n3. 不触发时限更新逻辑", "P0", "是", "反向验证场景"],
    
    ["TC-003", "跨专业关联触发", "主单子单时限比较逻辑", "主单时限 < 子单时限，需要更新",
     "1. 系统已配置集客和家宽专业规则\n2. 关联追单功能已启用",
     "1. 登录系统，创建主事件（集客专业）\n2. 设置主事件故障处理时限=30 分钟\n3. 创建子事件（家宽专业）\n4. 设置子事件故障处理时限=60 分钟\n5. 触发关联追单操作",
     "主事件：时限=30 分钟\n子事件：时限=60 分钟",
     "1. 系统判断主单处理时限 (30 分钟) < 子单处理时限 (60 分钟)=True\n2. 按子单处理时限 + 主单派发时间重新计算\n3. 更新所有相关时限字段", "P1", "是", "边界值场景"],
    
    ["TC-004", "跨专业关联触发", "主单子单时限比较逻辑", "主单时限 >= 子单时限，无需更新",
     "1. 系统已配置集客和家宽专业规则\n2. 关联追单功能已启用",
     "1. 登录系统，创建主事件（集客专业）\n2. 设置主事件故障处理时限=120 分钟\n3. 创建子事件（家宽专业）\n4. 设置子事件故障处理时限=60 分钟\n5. 触发关联追单操作",
     "主事件：时限=120 分钟\n子事件：时限=60 分钟",
     "1. 系统判断主单处理时限 (120 分钟) >= 子单处理时限 (60 分钟)=True\n2. 不触发时限更新逻辑\n3. 不触发驳回/撤销挂起操作", "P1", "是", "正常场景"],
    
    ["TC-005", "跨专业关联触发", "相同专业判断逻辑", "主事件和子事件专业相同，不触发跨专业逻辑",
     "1. 系统已配置集客和家宽专业规则\n2. 关联追单功能已启用",
     "1. 登录系统，创建主事件（集客专业）\n2. 设置主事件故障处理时限=60 分钟\n3. 创建子事件（集客专业，相同专业）\n4. 设置子事件故障处理时限=90 分钟\n5. 触发关联追单操作",
     "主事件：专业=集客，时限=60 分钟\n子事件：专业=集客，时限=90 分钟",
     "1. 系统判断主单专业与子单专业一致\n2. 不触发跨专业关联事件派单逻辑\n3. 不更新处理时限", "P1", "是", "异常场景"],
    
    ["TC-006", "跨专业关联触发", "非集客家宽专业判断", "非集客和家宽专业，不触发关联追单",
     "1. 系统已配置集客和家宽专业规则\n2. 关联追单功能已启用",
     "1. 登录系统，创建主事件（政企专业）\n2. 设置主事件故障处理时限=60 分钟\n3. 创建子事件（传输专业）\n4. 设置子事件故障处理时限=90 分钟\n5. 触发关联追单操作",
     "主事件：专业=政企，时限=60 分钟\n子事件：专业=传输，时限=90 分钟",
     "1. 系统判断主单专业不是集客或家宽\n2. 不触发关联追单逻辑\n3. 不更新处理时限", "P1", "是", "异常场景"],
    
    ["TC-007", "跨专业关联触发", "关联追单失败处理", "关联追单失败，不触发时限更新",
     "1. 系统已配置集客和家宽专业规则\n2. 关联追单功能已启用",
     "1. 登录系统，创建跨专业关联事件（集客 + 家宽）\n2. 模拟关联追单服务异常\n3. 触发关联追单操作",
     "主事件：专业=集客，时限=60 分钟\n子事件：专业=家宽，时限=120 分钟",
     "1. 关联追单失败\n2. 不触发时限更新逻辑\n3. 不执行任何后续操作", "P2", "否", "异常处理"],
    
    # ========== 模块二：延期/挂起状态处理 ==========
    ["TC-008", "延期挂起处理", "延期待审批状态驳回", "延期待审批状态下，时限变更自动驳回延期申请",
     "1. 系统已配置集客和家宽专业规则\n2. 关联追单功能已启用\n3. 任务处于延期待审批状态",
     "1. 登录系统，创建跨专业关联事件（集客 + 家宽）\n2. 任务处于延期待审批状态\n3. 触发关联追单成功，时限需要更新",
     "任务状态：延期待审批\n主单时限=30 分钟，子单时限=60 分钟",
     "1. 自动驳回延期申请\n2. 驳回原因：'影响【当前子事件的专业】业务，需按业务要求时限处理。'\n3. 更新任务处理时限\n4. 调度作战室留痕", "P0", "是", "核心反向场景"],
    
    ["TC-009", "延期挂起处理", "挂起待审批状态驳回", "挂起待审批状态下，时限变更自动驳回挂起申请",
     "1. 系统已配置集客和家宽专业规则\n2. 关联追单功能已启用\n3. 任务处于挂起待审批状态",
     "1. 登录系统，创建跨专业关联事件（集客 + 家宽）\n2. 任务处于挂起待审批状态\n3. 触发关联追单成功，时限需要更新",
     "任务状态：挂起待审批\n主单时限=30 分钟，子单时限=60 分钟",
     "1. 自动驳回挂起申请\n2. 驳回原因：'影响【当前子事件的专业】业务，需按业务要求时限处理。'\n3. 更新任务处理时限\n4. 调度作战室留痕", "P0", "是", "核心反向场景"],
    
    ["TC-010", "延期挂起处理", "挂起状态撤销回退个人", "挂起状态下，时限变更撤销挂起并回退到原处理人",
     "1. 系统已配置集客和家宽专业规则\n2. 关联追单功能已启用\n3. 任务处于挂起状态，由张三处理",
     "1. 登录系统，创建跨专业关联事件（集客 + 家宽）\n2. 任务处于挂起状态，由张三处理\n3. 触发关联追单成功，时限需要更新",
     "任务状态：挂起\n处理人：张三（电话：13800138000）\n主单时限=30 分钟，子单时限=60 分钟",
     "1. 取消挂起状态\n2. 任务当前处理人变更为原挂起申请的人员（张三）\n3. 调度作战室留痕：'撤销任务挂起，任务指派给【张三（13800138000）】'\n4. 更新处理时限", "P0", "是", "核心反向场景"],
    
    ["TC-011", "延期挂起处理", "个人分支回退逻辑", "挂起状态为个人分支，撤销后回退到个人",
     "1. 系统已配置集客和家宽专业规则\n2. 关联追单功能已启用\n3. 任务挂起类型为个人分支",
     "1. 登录系统，创建跨专业关联事件\n2. 任务处于挂起状态，类型为个人分支\n3. 触发关联追单成功",
     "任务挂起类型：个人分支\n主单时限=30 分钟，子单时限=60 分钟",
     "1. 撤销挂起状态\n2. 回退到个人（原处理人）", "P1", "是", "分支逻辑"],
    
    ["TC-012", "延期挂起处理", "同组处理回退逻辑", "挂起状态为同组处理，撤销后回退到组",
     "1. 系统已配置集客和家宽专业规则\n2. 关联追单功能已启用\n3. 任务挂起类型为同组处理",
     "1. 登录系统，创建跨专业关联事件\n2. 任务处于挂起状态，类型为同组处理\n3. 触发关联追单成功",
     "任务挂起类型：同组处理\n主单时限=30 分钟，子单时限=60 分钟",
     "1. 撤销挂起状态\n2. 回退到组（原处理组）", "P1", "是", "分支逻辑"],
    
    # ========== 模块三：IVR 通知功能 ==========
    ["TC-013", "IVR 通知功能", "个人 IVR 通知", "通知任务当前处理人（个人）",
     "1. 系统已配置集客和家宽专业规则\n2. IVR 通知功能已启用",
     "1. 登录系统，创建跨专业关联事件，时限需要更新\n2. 当前处理人为个人（张三）",
     "任务处理人：张三，电话：13800138000\n变更后的处理时限：120 分钟",
     "1. IVR 自动拨打张三电话\n2. 语音内容：'任务号【XXX】，任务名称【XXX】，影响【家宽】业务，变更处理时限为【120 分钟】，请及时处理。'", "P1", "是", "IVR 通知"],
    
    ["TC-014", "IVR 通知功能", "组 IVR 排班表通知", "通知当前处理组（按 IVR 排班表）",
     "1. 系统已配置集客和家宽专业规则\n2. IVR 通知功能已启用\n3. 当前处理组有 IVR 排班表",
     "1. 登录系统，创建跨专业关联事件，时限需要更新\n2. 当前处理人为组（如：集客支撑组）",
     "任务处理组：集客支撑组\n变更后的处理时限：120 分钟",
     "1. 系统根据对应组的 IVR 排班表进行通知\n2. IVR 内容包含完整信息", "P1", "是", "IVR 通知"],
    
    ["TC-015", "IVR 通知功能", "IVR 电话为空异常处理", "任务当前处理人电话为空，IVR 通知失败",
     "1. 系统已配置集客和家宽专业规则\n2. IVR 通知功能已启用",
     "1. 登录系统，创建跨专业关联事件，时限需要更新\n2. 当前处理人电话为空或无效",
     "任务处理人：张三，电话：空\n变更后的处理时限：120 分钟",
     "1. IVR 通知失败\n2. 系统记录告警日志\n3. 触发重试机制或人工干预流程", "P2", "否", "异常处理"],
    
    # ========== 模块四：调度作战室留痕 ==========
    ["TC-016", "调度作战室留痕", "时限变更留痕记录", "调度作战室留痕内容完整性验证",
     "1. 系统已配置集客和家宽专业规则\n2. 调度作战室留痕功能已启用",
     "1. 登录系统，创建跨专业关联事件，时限需要更新\n2. 触发关联追单成功",
     "原任务处理时限：60 分钟\n变更后任务处理时限：120 分钟",
     "1. 留痕标题='调度指令'\n2. 留痕内容包含：'影响【当前子事件的专业】业务，任务处理时限由【60 分钟】变更为【120 分钟】'", "P1", "是", "日志留痕"],
    
    # ========== 模块五：状态流转控制 ==========
    ["TC-017", "状态流转控制", "正常处理状态更新", "正常处理状态下，时限变更仅更新不驳回",
     "1. 系统已配置集客和家宽专业规则\n2. 任务状态=正常处理",
     "1. 登录系统，创建跨专业关联事件，任务处于正常处理状态\n2. 触发关联追单成功，时限需要更新",
     "任务状态：正常处理\n主单时限=30 分钟，子单时限=60 分钟",
     "1. 仅更新处理时限\n2. 不触发驳回操作\n3. 不撤销挂起", "P1", "是", "状态流转"],
    
    # ========== 模块六：边界值测试 ==========
    ["TC-018", "边界值测试", "时限相等边界测试", "主单时限 = 子单时限，边界值相等",
     "1. 系统已配置集客和家宽专业规则\n2. 关联追单功能已启用",
     "1. 登录系统，创建跨专业关联事件（集客 + 家宽）\n2. 主单时限=60 分钟，子单时限=60 分钟\n3. 触发关联追单成功",
     "主单时限：60 分钟\n子单时限：60 分钟",
     "1. 系统判断主单处理时限 (60 分钟) < 子单处理时限 (60 分钟)=False\n2. 不触发时限更新", "P1", "是", "边界值"],
    
    ["TC-019", "边界值测试", "时限临界值测试", "主单时限 = 子单时限 -1，边界值临界",
     "1. 系统已配置集客和家宽专业规则\n2. 关联追单功能已启用",
     "1. 登录系统，创建跨专业关联事件（集客 + 家宽）\n2. 主单时限=59 分钟，子单时限=60 分钟\n3. 触发关联追单成功",
     "主单时限：59 分钟\n子单时限：60 分钟",
     "1. 系统判断主单处理时限 (59 分钟) < 子单处理时限 (60 分钟)=True\n2. 触发时限更新", "P1", "是", "边界值"],
    
    # ========== 模块七：数据完整性验证 ==========
    ["TC-020", "数据完整性验证", "所有时限字段更新验证", "更新所有时限字段完整性验证",
     "1. 系统已配置集客和家宽专业规则\n2. 关联追单功能已启用",
     "1. 登录系统，创建跨专业关联事件\n2. 触发关联追单成功，时限需要更新",
     "主事件：专业=集客，时限=30 分钟\n子事件：专业=家宽，时限=60 分钟",
     "1. 故障受理时限已更新\n2. 故障处理时限已更新\n3. T1 受理时限已更新\n4. T1 处理时限已更新\n5. T2 受理时限已更新\n6. T2 处理时限已更新", "P0", "是", "数据完整性"],
]

# 创建 Excel 文件
output_file = "/Users/linziwang/PycharmProjects/wordToWord/docs/测试/跨专业派单时限优化_测试用例.xlsx"
wb = Workbook()

# 创建测试用例总览工作表
ws_test_cases = wb.active
ws_test_cases.title = "测试用例总览"

# 定义样式
header_font = Font(name='微软雅黑', size=10, bold=True, color="FFFFFF")
normal_font = Font(name='微软雅黑', size=9)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
alt_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color="000000"),
    right=Side(style='thin', color="000000"),
    top=Side(style='thin', color="000000"),
    bottom=Side(style='thin', color="000000")
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

# 写入表头
headers = ["用例编号", "测试模块", "功能点", "用例标题", "前置条件", "测试步骤", "输入数据", "预期结果", "优先级", "是否自动化", "备注"]
for col, header in enumerate(headers, 1):
    cell = ws_test_cases.cell(row=1, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

# 写入数据行
for row_idx, case in enumerate(test_cases, 2):
    for col_idx, value in enumerate(case, 1):
        cell = ws_test_cases.cell(row=row_idx, column=col_idx)
        cell.value = value
        
        if row_idx % 2 == 0:
            cell.fill = alt_row_fill
        else:
            cell.fill = row_fill
        
        cell.border = thin_border
        cell.font = normal_font
        
        if col_idx <= 10:
            cell.alignment = center_align
        else:
            cell.alignment = left_align

# 调整列宽 - 根据内容长度设置合适的宽度
column_widths = {
    'A': 12,   # 用例编号
    'B': 15,   # 测试模块
    'C': 18,   # 功能点
    'D': 35,   # 用例标题
    'E': 40,   # 前置条件
    'F': 50,   # 测试步骤
    'G': 40,   # 输入数据
    'H': 85,   # 预期结果
    'I': 10,   # 优先级
    'J': 12,   # 是否自动化
    'K': 15    # 备注
}

for col, width in column_widths.items():
    ws_test_cases.column_dimensions[col].width = width

# 创建统计信息工作表
ws_stats = wb.create_sheet("统计信息")
stats_data = [
    ["测试类型", "用例数量", "占比"],
    ["正向测试", 4, "20%"],
    ["反向测试", 5, "25%"],
    ["异常测试", 4, "20%"],
    ["边界值测试", 2, "10%"],
    ["数据完整性测试", 1, "5%"],
    ["IVR 通知测试", 2, "10%"],
    ["日志留痕测试", 1, "5%"],
    ["状态流转测试", 1, "5%"],
    ["总计", 20, "100%"]
]

for row_idx, header in enumerate(["测试类型", "用例数量", "占比"], 1):
    cell = ws_stats.cell(row=row_idx, column=row_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

for row_idx, stats in enumerate(stats_data, 2):
    for col_idx, value in enumerate(stats, 1):
        cell = ws_stats.cell(row=row_idx, column=col_idx)
        cell.value = value
        
        if row_idx % 2 == 0:
            cell.fill = alt_row_fill
        else:
            cell.fill = row_fill
        
        cell.border = thin_border
        if col_idx <= 2:
            cell.alignment = center_align
    
    ws_stats.column_dimensions[chr(64 + col_idx)].width = 10 if col_idx == 2 else (8 if col_idx == 3 else 15)

# 创建优先级说明工作表
ws_priority = wb.create_sheet("优先级说明")
priority_data = [
    ["P0", "高优先级 - 核心功能，必须测试"],
    ["P1", "中优先级 - 重要功能，建议测试"],
    ["P2", "低优先级 - 异常场景，可选测试"]
]

for row_idx, header in enumerate(["优先级", "说明"], 1):
    cell = ws_priority.cell(row=row_idx, column=row_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

for row_idx, priority in enumerate(priority_data, 2):
    for col_idx, value in enumerate(priority, 1):
        cell = ws_priority.cell(row=row_idx, column=col_idx)
        cell.value = value
        
        if row_idx % 2 == 0:
            cell.fill = alt_row_fill
        else:
            cell.fill = row_fill
        
        cell.border = thin_border
    
    ws_priority.column_dimensions[chr(64 + col_idx)].width = 10 if col_idx == 1 else 50

# 创建测试模块说明工作表
ws_modules = wb.create_sheet("测试模块说明")
modules_data = [
    ["模块编号", "测试模块", "功能点", "用例数量", "说明"],
    ["M01", "跨专业关联触发", "关联追单触发条件验证", 7, "验证集客和家宽跨专业场景下的关联追单触发逻辑"],
    ["M02", "延期挂起处理", "延期待审批/挂起状态驳回", 5, "验证任务处于延期或挂起状态时的自动处理逻辑"],
    ["M03", "IVR 通知功能", "个人/组 IVR 排班表通知", 3, "验证 IVR 电话通知功能的正确性"],
    ["M04", "调度作战室留痕", "时限变更留痕记录", 1, "验证调度作战室的留痕功能"],
    ["M05", "状态流转控制", "正常处理状态更新", 1, "验证任务状态流转的正确性"],
    ["M06", "边界值测试", "时限相等/临界值测试", 2, "验证边界条件下的系统行为"],
    ["M07", "数据完整性验证", "所有时限字段更新验证", 1, "验证所有相关时限字段的完整更新"]
]

for row_idx, header in enumerate(["模块编号", "测试模块", "功能点", "用例数量", "说明"], 1):
    cell = ws_modules.cell(row=row_idx, column=row_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

for row_idx, module in enumerate(modules_data, 2):
    for col_idx, value in enumerate(module, 1):
        cell = ws_modules.cell(row=row_idx, column=col_idx)
        cell.value = value
        
        if row_idx % 2 == 0:
            cell.fill = alt_row_fill
        else:
            cell.fill = row_fill
        
        cell.border = thin_border
    
    ws_modules.column_dimensions[chr(64 + col_idx)].width = 12 if col_idx <= 3 else (8 if col_idx == 4 else 50)

# 保存文件
wb.save(output_file)
print(f"✅ Excel 文件已生成：{output_file}")

# 打印统计信息
print("\n" + "="*80)
print("测试用例设计完成！")
print("="*80)

# 按模块统计
module_counts = {}
for case in test_cases:
    module = case[1]  # 测试模块列
    module_counts[module] = module_counts.get(module, 0) + 1

print("\n各测试模块用例分布:")
for module, count in sorted(module_counts.items()):
    print(f"  - {module}: {count} 个用例")

print(f"\n总用例数：{len(test_cases)} 个")
print(f"自动化测试覆盖率：{sum(1 for c in test_cases if c[9] == '是')}/{len(test_cases)} = {sum(1 for c in test_cases if c[9] == '是')/len(test_cases)*100:.1f}%")

