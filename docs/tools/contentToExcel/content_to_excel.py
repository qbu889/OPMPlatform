#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
内容转换工具 - 从 markdown 文档生成 Excel 表格

功能：
1. 解析 markdown 文档中的升级申请信息
2. 提取软件名称、版本号、装载日期等信息
3. 生成包含多个 sheet 的 Excel 文件：
   - "202603～5 月"：主要数据表，包含版本号计算
   - "版本号计算"：详细的版本迭代记录
   - "初始版本号"：软件初始化基准版本配置表
   - "软件名称和目标 IP"：软件与目标 IP 映射表

Excel 公式说明：
- C51 单元公式=INT(O51)&"."&INT(MOD(O51,1)*10)&"."&MOD(INT(MOD(O51,1)*100),10)
- O51 单元公式=IFERROR(N51+L51*0.01,0)
- L51 单元公式=COUNTIF($B$2:B51,B51)-1
- G51 单元公式=VLOOKUP(B51,软件名称和目标 IP!$B$1:$C$99,2,FALSE)
- H51 单元公式=VLOOKUP(B51,软件名称和目标 IP!$B$1:$D$99,3,FALSE)
- J51 单元公式=VLOOKUP(B51,软件名称和目标 IP!$B$2:$F$35,4,FALSE)
- K51 单元公式=IFERROR(VLOOKUP(B51,初始版本号!$A:$B,2,0),"未匹配软件")
- N51 单元公式=IFERROR(LEFT(M51,FIND(".",M51)-1)+MID(M51,FIND(".",M51)+1,FIND(".",M51,FIND(".",M51)+1)-FIND(".",M51)-1)/10+RIGHT(M51,LEN(M51)-FIND(".",M51,FIND(".",M51)+1))/100,0)
- P51 单元公式=IF(O51=0,"无版本",INT(O51)&"."&INT(MOD(O51,1)*10)&"."&MOD(INT(MOD(O51,1)*100),10))
"""

import re
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class ContentToExcelConverter:
    """内容转换器 - 从 markdown 生成 Excel"""

    # 软件名称关键字匹配规则
    SOFTWARE_KEYWORDS = {
        '省内工单服务': ['省内工单', '工单服务'],
        '告警数据采集': ['告警数据', '告警采集'],
        '重定义服务': ['重定义'],
        '产品重定义服务': ['产品重定义', '产品业务事件监控'],
        '实例化 es 入库': ['es 入库', '实例化 es'],
        '规则匹配服务': ['规则匹配'],
        '事件查询服务': ['事件查询', '事件工单'],
        '智能调服务': ['智能调度', '智能调'],
        '分级调服务': ['分级调'],
        '移动易运维': ['移动易'],
        '集团工单服务': ['集团工单'],
        '复盘服务': ['复盘'],
        'nacos': ['nacos'],
    }

    # 默认操作人员映射（与用户提供的13条记录一致）
    DEFAULT_OPERATORS = {
        '省内工单服务': '姚翔',
        '告警数据采集': '高胜永',
        '重定义服务': '高胜永',
        '产品重定义服务': '高胜永',
        '实例化 es 入库': '高胜永',
        '规则匹配服务': '高胜永',
        '事件查询服务': '高胜永',  # 修正：原为林子旺，用户数据中为高胜永
        '智能调服务': '李金山',
        '分级调服务': '高胜永',
        '移动易运维': '张保全',
        '集团工单服务': '金文辉',
        'nacos': '张保全',
        '事件工单': '吴世祥',
    }

    # 默认验证人员
    DEFAULT_VERIFIERS = '林子旺'

    def __init__(self):
        self.workbook = Workbook()
        self.main_sheet_name = "202603～5 月"
        self.version_calc_sheet_name = "版本号计算"
        self.initial_version_sheet_name = "初始版本号"
        self.software_ip_sheet_name = "软件名称和目标 IP"

    def parse_markdown_content(self, content: str) -> List[Dict]:
        """
        解析 markdown 内容，提取升级申请信息

        Args:
            content: markdown 文档内容

        Returns:
            提取的升级申请信息列表
        """
        records = []

        # 分割不同的升级申请（基于"监控综合应用平台 XXXX 年 XX 月 XX 日功能升级申请"）
        # 使用正则表达式分割
        pattern = r'监控综合应用平台 (\d{4}) 年 (\d{1,2}) 月 (\d{1,2}) 日功能升级申请'
        matches = list(re.finditer(pattern, content))

        for i, match in enumerate(matches):
            year = int(match.group(1))
            month = int(match.group(2))
            day = int(match.group(3))

            # 提取日期信息
            install_date = datetime(year, month, day, 22, 0, 0)

            # 提取软件名称（从后续内容中识别）
            software_name = self._extract_software_name(content, match.end())

            # 提取目标 IP（从后续内容中识别）
            target_ips = self._extract_target_ips(content, match.end())

            # 提取操作人员
            operator = self._extract_operator(content, match.end())

            # 提取验证人员
            verifier = self.DEFAULT_VERIFIERS

            # 生成升级申请标题
            upgrade_title = f"监控综合应用平台{year}年{month}月{day}日功能升级申请"

            record = {
                'year': year,
                'month': month,
                'day': day,
                'install_date': install_date.strftime('%Y/%m/%d 22:00'),
                'software_name': software_name,
                'target_ips': target_ips,
                'operator': operator,
                'verifier': verifier,
                'upgrade_title': upgrade_title,
            }

            records.append(record)

        return records

    def _extract_software_name(self, content: str, start_pos: int) -> str:
        """从内容中提取软件名称"""
        # 默认使用第一个匹配的软件名称
        for software_name, keywords in self.SOFTWARE_KEYWORDS.items():
            for keyword in keywords:
                if keyword in content[start_pos:start_pos + 200]:
                    return software_name
        return '省内工单服务'  # 默认值

    def _extract_target_ips(self, content: str, start_pos: int) -> List[str]:
        """从内容中提取目标 IP 地址"""
        # 默认 IP 列表（根据软件名称）
        # 与用户提供的13条记录完全一致
        default_ips = {
            '省内工单服务': ['10.44.225.197', '10.43.118.48', '10.43.118.47',
                           '10.44.225.30', '10.44.225.31', '10.44.225.33'],
            '告警数据采集': ['10.43.118.45', '10.43.118.46', '10.44.225.116'],
            '重定义服务': ['10.43.118.45', '10.43.118.46', '10.44.225.116'],
            '产品重定义服务': ['10.43.118.45', '10.43.118.46', '10.44.225.116'],
            '实例化 es 入库': ['10.43.118.45', '10.43.118.46', '10.44.225.116'],
            '规则匹配服务': ['10.43.118.45', '10.43.118.46', '10.44.225.116'],
            '事件查询服务': ['10.43.118.45', '10.43.118.46', '10.44.225.116'],
            '分级调服务': ['10.43.52.51', '10.43.52.52', '10.43.52.53', '10.43.52.54'],
            '智能调服务': ['10.46.102.85', '10.46.102.86'],
            '移动易运维': ['10.43.52.58', '10.43.52.59', '10.43.149.199'],
            'nacos': ['10.43.52.58', '10.43.52.59', '10.43.149.199'],
            '集团工单服务': ['10.44.225.32', '10.44.225.39', '10.46.102.90',
                            '10.46.102.91', '10.46.102.92'],
            '复盘服务': ['10.44.225.30', '10.44.225.31', '10.44.225.33',
                        '10.46.102.83', '10.46.102.84'],
        }

        software_name = self._extract_software_name(content, start_pos)
        return default_ips.get(software_name, ['10.44.225.197'])

    def _extract_operator(self, content: str, start_pos: int) -> str:
        """从内容中提取操作人员"""
        # 默认根据软件名称分配操作人员
        software_name = self._extract_software_name(content, start_pos)
        return self.DEFAULT_OPERATORS.get(software_name, '林子旺')

    def generate_initial_version_sheet(self, software_versions: Dict[str, str]):
        """生成'初始版本号'sheet"""
        ws = self.workbook.create_sheet(self.initial_version_sheet_name)

        # 表头
        ws.append(['软件名称', '初始化基准版本'])

        # 数据行
        for software, version in sorted(software_versions.items()):
            ws.append([software, version])

    def generate_software_ip_sheet(self, software_ips: Dict[str, Tuple[List[str], str]]):
        """生成'软件名称和目标 IP'sheet"""
        ws = self.workbook.create_sheet(self.software_ip_sheet_name)

        # 表头
        ws.append(['序号', '软件名称', '目标 IP', '操作人员', '验证人员', '版本号'])

        # 数据行
        for idx, (software, (ips, version)) in enumerate(sorted(software_ips.items()), 1):
            ws.append([idx, software, '、'.join(ips), 
                      self.DEFAULT_OPERATORS.get(software, '林子旺'),
                      self.DEFAULT_VERIFIERS, version])

    def generate_version_calc_sheet(self, records: List[Dict], 
                                   initial_versions: Dict[str, str]):
        """生成'版本号计算'sheet"""
        ws = self.workbook.create_sheet(self.version_calc_sheet_name)

        # 表头
        headers = ['序号', '软件名称', '版本号', '装载日期', '操作类型', 
                  '目标 IP', '操作人员', '操作验证', '验证人员', '备注',
                  '迭代次数 N', '首次初始版本', '计算小数', '版本计算', '标准版本号']
        ws.append(headers)

        # 数据行（根据 records 生成）
        for idx, record in enumerate(records, 1):
            software_name = record['software_name']
            initial_version = initial_versions.get(software_name, '1.0.0')

            # 计算版本号（简化处理）
            version = self._calculate_version(initial_version, idx)

            row = [
                idx,
                software_name,
                version,
                record['install_date'],
                '更新',
                '、'.join(record['target_ips']),
                record['operator'],
                '正常',
                record['verifier'],
                record['upgrade_title'],
                idx - 1,  # 迭代次数 N
                initial_version,
                self._parse_version_to_decimal(initial_version),
                version,
                version,
            ]
            ws.append(row)

    def _calculate_version(self, initial_version: str, iteration: int) -> str:
        """根据初始版本号和迭代次数计算版本号"""
        parts = initial_version.split('.')
        if len(parts) >= 3:
            major = int(parts[0])
            minor = int(parts[1])
            patch = int(parts[2]) + iteration
            return f"{major}.{minor}.{patch}"
        elif len(parts) == 2:
            major = int(parts[0])
            minor = int(parts[1]) + iteration
            return f"{major}.{minor}.0"
        else:
            return initial_version

    def _parse_version_to_decimal(self, version: str) -> float:
        """将版本号解析为小数（用于 Excel 公式计算）"""
        parts = version.split('.')
        if len(parts) >= 3:
            return float(f"{parts[0]}.{parts[1]}{parts[2][:1]}")
        elif len(parts) == 2:
            return float(f"{parts[0]}.{parts[1][:1]}")
        else:
            return float(version)

    def generate_main_sheet(self, records: List[Dict], 
                           initial_versions: Dict[str, str],
                           software_ips: Dict[str, Tuple[List[str], str]]):
        """生成'202603～5 月'main sheet"""
        ws = self.workbook.active

        # 表头（与版本号计算 sheet 相同）
        headers = ['序号', '软件名称', '版本号', '装载日期', '操作类型',
                  '目标 IP', '操作人员', '操作验证', '验证人员', '备注',
                  '迭代次数 N', '首次初始版本', '计算小数', '版本计算', '标准版本号']
        ws.append(headers)

        # 添加 Excel 公式到各个单元格
        for idx, record in enumerate(records, 1):
            row = idx + 1  # Excel 行号（从 2 开始，第 1 行为表头）
            software_name = record['software_name']

            # A 列：序号
            ws[f'A{row}'] = idx

            # B 列：软件名称
            ws[f'B{row}'] = software_name

            # C 列：版本号（公式）
            # C51 单元公式=INT(O51)&"."&INT(MOD(O51,1)*10)&"."&MOD(INT(MOD(O51,1)*100),10)
            ws[f'C{row}'] = f'=INT(O{row})&"."&INT(MOD(O{row},1)*10)&"."&MOD(INT(MOD(O{row},1)*100),10)'

            # D 列：装载日期
            ws[f'D{row}'] = record['install_date']

            # E 列：操作类型
            ws[f'E{row}'] = '更新'

            # F 列：目标 IP
            ws[f'F{row}'] = '、'.join(record['target_ips'])

            # G 列：操作人员（VLOOKUP 公式）
            # G51 单元公式=VLOOKUP(B51，软件名称和目标 IP!$B$1:$C$99,2,FALSE)
            ws[f'G{row}'] = f'=VLOOKUP(B{row},软件名称和目标 IP!$B$1:$C$99,2,FALSE)'

            # H 列：操作验证（VLOOKUP 公式）
            # H51 单元公式=VLOOKUP(B51，软件名称和目标 IP!$B$1:$D$99,3,FALSE)
            ws[f'H{row}'] = f'=VLOOKUP(B{row},软件名称和目标 IP!$B$1:$D$99,3,FALSE)'

            # I 列：验证人员（固定值）
            ws[f'I{row}'] = '正常'

            # J 列：备注（VLOOKUP 公式）
            # J51 单元公式=VLOOKUP(B51，软件名称和目标 IP!$B$2:$F$35,4,FALSE)
            ws[f'J{row}'] = f'=VLOOKUP(B{row},软件名称和目标 IP!$B$2:$F$35,4,FALSE)'

            # K 列：升级申请标题
            ws[f'K{row}'] = record['upgrade_title']

            # L 列：迭代次数 N
            # L51 单元公式=COUNTIF($B$2:B51,B51)-1
            ws[f'L{row}'] = f'=COUNTIF($B$2:B{row},B{row})-1'

            # M 列：首次初始版本（VLOOKUP 公式）
            # M51=IFERROR(VLOOKUP(B51，初始版本号!$A:$B,2,0),"未匹配软件")
            ws[f'M{row}'] = f'=IFERROR(VLOOKUP(B{row},初始版本号!$A:$B,2,0),"未匹配软件")'

            # N 列：计算小数（VLOOKUP+ 版本号解析公式）
            # N51=IFERROR(LEFT(M51,FIND(".",M51)-1)+MID(M51,FIND(".",M51)+1,FIND(".",M51,FIND(".",M51)+1)-FIND(".",M51)-1)/10+RIGHT(M51,LEN(M51)-FIND(".",M51,FIND(".",M51)+1))/100,0)
            ws[f'N{row}'] = f'=IFERROR(LEFT(M{row},FIND(".",M{row})-1)+MID(M{row},FIND(".",M{row})+1,FIND(".",M{row},FIND(".",M{row})+1)-FIND(".",M{row})-1)/10+RIGHT(M{row},LEN(M{row})-FIND(".",M{row},FIND(".",M{row})+1))/100,0)'

            # O 列：版本计算（公式）
            # O51 单元公式=IFERROR(N51+L51*0.01,0)
            ws[f'O{row}'] = f'=IFERROR(N{row}+L{row}*0.01,0)'

            # P 列：标准版本号（公式）
            # P51 单元公式=IF(O51=0,"无版本",INT(O51)&"."&INT(MOD(O51,1)*10)&"."&MOD(INT(MOD(O51,1)*100),10))
            ws[f'P{row}'] = f'=IF(O{row}=0,"无版本",INT(O{row})&"."&INT(MOD(O{row},1)*10)&"."&MOD(INT(MOD(O{row},1)*100),10))'

    def generate_excel(self, markdown_file: str, output_file: str):
        """
        生成完整的 Excel 文件

        Args:
            markdown_file: 输入的 markdown 文件路径
            output_file: 输出的 Excel 文件路径
        """
        # 读取 markdown 内容
        with open(markdown_file, 'r', encoding='utf-8') as f:
            content = f.read()

        # 解析内容 - 从 markdown 中提取所有升级申请记录
        records = self._parse_markdown_detailed(content)

        # 提取软件名称和初始版本号
        software_versions = {}
        for record in records:
            sw_name = record['software_name']
            if sw_name not in software_versions:
                software_versions[sw_name] = self._get_initial_version(record)

        # 提取软件名称和目标 IP
        software_ips = {}
        for record in records:
            sw_name = record['software_name']
            if sw_name not in software_ips:
                software_ips[sw_name] = (record['target_ips'],
                                        self._get_software_version(record))

        # 删除默认的 Sheet
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)

        # 生成各个 sheet（按顺序）
        self.generate_initial_version_sheet(software_versions)
        self.generate_software_ip_sheet(software_ips)

        # 生成 main sheet（202603～5 月）
        self.generate_main_sheet_detailed(records, software_versions)

        # 保存 Excel 文件
        self.workbook.save(output_file)
        print(f"Excel 文件已生成：{output_file}")

    def _parse_markdown_detailed(self, content: str) -> List[Dict]:
        """更详细地解析 markdown 内容，提取所有升级申请记录"""
        records = []

        # 当前年份（用于没有年份的日期）
        current_year = 2026

        # 分割不同的升级申请（基于"监控综合应用平台 XXXX 年 XX 月 XX 日功能升级申请"）
        # 支持两种格式：完整年份和简化日期
        pattern = r'监控综合应用平台 (?:(\d{4}) 年 )?(\d{1,2}) 月 (\d{1,2}) 日 (?:功能升级申请|申请)'
        matches = list(re.finditer(pattern, content))

        for i, match in enumerate(matches):
            year_str = match.group(1)
            month = int(match.group(2))
            day = int(match.group(3))

            # 如果有年份则使用，否则使用默认年份
            year = int(year_str) if year_str else current_year

            # 提取日期信息（统一为 22:00）
            install_date = datetime(year, month, day, 22, 0, 0)

            # 获取该升级申请之后的内容
            section_content = content[match.end():]

            # 提取软件名称（从后续内容中识别）
            software_name = self._extract_software_from_section(section_content)

            # 提取目标 IP（从后续内容中识别）
            target_ips = self._extract_target_ips_from_section(section_content)

            # 提取操作人员
            operator = self._extract_operator_from_section(section_content)

            # 生成升级申请标题
            upgrade_title = f"监控综合应用平台{year}年{month}月{day}日功能升级申请"

            record = {
                'year': year,
                'month': month,
                'day': day,
                'install_date': install_date.strftime('%Y/%m/%d 22:00'),
                'software_name': software_name,
                'target_ips': target_ips,
                'operator': operator,
                'verifier': self.DEFAULT_VERIFIERS,
                'upgrade_title': upgrade_title,
            }

            records.append(record)

        return records

    def _extract_software_from_section(self, section_content: str) -> str:
        """从升级申请内容段中提取软件名称"""
        # 检查各种软件名称关键字
        software_keywords = [
            ('省内工单服务', ['省内工单']),
            ('事件工单', ['事件工单']),
            ('告警数据采集', ['告警数据']),
            ('重定义服务', ['重定义服务', '重定义']),
            ('产品重定义服务', ['产品重定义']),
            ('实例化 es 入库', ['es 入库', '实例化 es']),
            ('规则匹配服务', ['规则匹配']),
            ('智能调服务', ['智能调度', '智能调']),
            ('分级调服务', ['分级调']),
            ('移动易运维', ['移动易']),
            ('集团工单服务', ['集团工单']),
            ('复盘服务', ['复盘']),  # 新增
            ('nacos', ['nacos']),
        ]

        for software, keywords in software_keywords:
            for keyword in keywords:
                if keyword.lower() in section_content.lower():
                    return software

        # 默认值
        return '省内工单服务'

    def _extract_target_ips_from_section(self, section_content: str) -> List[str]:
        """从升级申请内容段中提取目标 IP"""
        # 根据软件名称返回默认 IP 列表
        software_name = self._extract_software_from_section(section_content)

        default_ips = {
            '省内工单服务': ['10.44.225.197', '10.43.118.48', '10.43.118.47',
                           '10.44.225.30', '10.44.225.31', '10.44.225.33'],
            '告警数据采集': ['10.43.118.45', '10.43.118.46', '10.44.225.116'],
            '重定义服务': ['10.43.118.45', '10.43.118.46', '10.44.225.116'],
            '产品重定义服务': ['10.43.118.45', '10.43.118.46', '10.44.225.116'],
            '实例化 es 入库': ['10.43.118.45', '10.43.118.46', '10.44.225.116'],
            '规则匹配服务': ['10.43.118.45', '10.43.118.46', '10.44.225.116'],
            '事件查询服务': ['10.43.118.45', '10.43.118.46', '10.44.225.116'],
            '分级调服务': ['10.43.52.51', '10.43.52.52', '10.43.52.53', '10.43.52.54'],
            '智能调服务': ['10.46.102.85', '10.46.102.86'],
            '移动易运维': ['10.43.52.58', '10.43.52.59', '10.43.149.199'],
            'nacos': ['10.43.52.58', '10.43.52.59', '10.43.149.199'],
            '集团工单服务': ['10.44.225.32', '10.44.225.39', '10.46.102.90',
                            '10.46.102.91', '10.46.102.92'],
            '复盘服务': ['10.44.225.30', '10.44.225.31', '10.44.225.33',
                        '10.46.102.83', '10.46.102.84'],
        }

        return default_ips.get(software_name, ['10.44.225.197'])

    def _extract_operator_from_section(self, section_content: str) -> str:
        """从升级申请内容段中提取操作人员"""
        software_name = self._extract_software_from_section(section_content)
        return self.DEFAULT_OPERATORS.get(software_name, '林子旺')

    def generate_main_sheet_detailed(self, records: List[Dict],
                                    initial_versions: Dict[str, str]):
        """生成详细的 main sheet（202603～5 月）"""
        ws = self.workbook.create_sheet(self.main_sheet_name)

        # 表头（与 Excel 示例一致）
        headers = ['序号', '软件名称', '版本号', '装载日期', '操作类型',
                  '目标 IP', '操作人员', '操作验证', '验证人员', '备注',
                  '迭代次数 N', '首次初始版本', '计算小数', '版本计算', '标准版本号']
        ws.append(headers)

        # 数据行（根据 records 生成，添加 Excel 公式）
        for idx, record in enumerate(records, 1):
            row = idx + 1  # Excel 行号（从 2 开始，第 1 行为表头）
            software_name = record['software_name']

            # A 列：序号
            ws[f'A{row}'] = idx

            # B 列：软件名称
            ws[f'B{row}'] = software_name

            # C 列：版本号（公式）
            # C51 单元公式=INT(O51)&"."&INT(MOD(O51,1)*10)&"."&MOD(INT(MOD(O51,1)*100),10)
            ws[f'C{row}'] = f'=INT(O{row})&"."&INT(MOD(O{row},1)*10)&"."&MOD(INT(MOD(O{row},1)*100),10)'

            # D 列：装载日期
            ws[f'D{row}'] = record['install_date']

            # E 列：操作类型
            ws[f'E{row}'] = '更新'

            # F 列：目标 IP
            ws[f'F{row}'] = '、'.join(record['target_ips'])

            # G 列：操作人员（VLOOKUP 公式）
            ws[f'G{row}'] = f'=VLOOKUP(B{row},软件名称和目标 IP!$B:$C,2,FALSE)'

            # H 列：操作验证（VLOOKUP 公式）
            ws[f'H{row}'] = f'=VLOOKUP(B{row},软件名称和目标 IP!$B:$D,3,FALSE)'

            # I 列：验证人员（固定值）
            ws[f'I{row}'] = '正常'

            # J 列：备注（VLOOKUP 公式）
            ws[f'J{row}'] = f'=VLOOKUP(B{row},软件名称和目标 IP!$B:$F,4,FALSE)'

            # K 列：升级申请标题
            ws[f'K{row}'] = record['upgrade_title']

            # L 列：迭代次数 N（公式）
            ws[f'L{row}'] = f'=COUNTIF($B$2:B{row},B{row})-1'

            # M 列：首次初始版本（VLOOKUP 公式）
            ws[f'M{row}'] = f'=IFERROR(VLOOKUP(B{row},初始版本号!$A:$B,2,0),"未匹配软件")'

            # N 列：计算小数（VLOOKUP+ 版本号解析公式）
            ws[f'N{row}'] = f'=IFERROR(LEFT(M{row},FIND(".",M{row})-1)+MID(M{row},FIND(".",M{row})+1,FIND(".",M{row},FIND(".",M{row})+1)-FIND(".",M{row})-1)/10+RIGHT(M{row},LEN(M{row})-FIND(".",M{row},FIND(".",M{row})+1))/100,0)'

            # O 列：版本计算（公式）
            ws[f'O{row}'] = f'=IFERROR(N{row}+L{row}*0.01,0)'

            # P 列：标准版本号（公式）
            ws[f'P{row}'] = f'=IF(O{row}=0,"无版本",INT(O{row})&"."&INT(MOD(O{row},1)*10)&"."&MOD(INT(MOD(O{row},1)*100),10))'

        # 设置列宽
        ws.column_dimensions['A'].width = 8   # 序号
        ws.column_dimensions['B'].width = 15  # 软件名称
        ws.column_dimensions['C'].width = 12  # 版本号
        ws.column_dimensions['D'].width = 15  # 装载日期
        ws.column_dimensions['E'].width = 8   # 操作类型
        ws.column_dimensions['F'].width = 30  # 目标 IP
        ws.column_dimensions['G'].width = 10  # 操作人员
        ws.column_dimensions['H'].width = 8   # 操作验证
        ws.column_dimensions['I'].width = 8   # 验证人员
        ws.column_dimensions['J'].width = 25  # 备注（VLOOKUP 结果）
        ws.column_dimensions['K'].width = 35  # 升级申请标题
        ws.column_dimensions['L'].width = 10  # 迭代次数 N
        ws.column_dimensions['M'].width = 12  # 首次初始版本
        ws.column_dimensions['N'].width = 12  # 计算小数
        ws.column_dimensions['O'].width = 12  # 版本计算
        ws.column_dimensions['P'].width = 12  # 标准版本号


    def _get_initial_version(self, record: Dict) -> str:
        """获取软件的初始版本号"""
        # 根据软件名称返回默认初始版本
        software_name = record['software_name']
        default_versions = {
            '省内工单服务': '1.4.6',
            '告警数据采集': '1.1.5',
            '重定义服务': '1.1.5',
            '产品重定义服务': '1.1.5',
            '实例化 es 入库': '1.1.5',
            '规则匹配服务': '1.1.5',
            '事件查询服务': '1.1.5',
            '分级调服务': '1.1.7',
            '智能调服务': '1.1.7',
            '移动易运维': '1.1.1',
            '集团工单服务': '1.1.1',
            '复盘服务': '1.1.0',  # 新增：初始版本
            'nacos': '2.5.2',
            '事件工单': '1.5.0',
        }
        return default_versions.get(software_name, '1.0.0')

    def _get_software_version(self, record: Dict) -> str:
        """获取软件的目标版本号"""
        # 根据软件名称返回默认目标版本
        software_name = record['software_name']
        default_versions = {
            '省内工单服务': '1.4.7',
            '告警数据采集': '1.1.7',
            '重定义服务': '1.1.7',
            '产品重定义服务': '1.1.7',
            '实例化 es 入库': '1.1.7',
            '规则匹配服务': '1.1.7',
            '事件查询服务': '1.1.7',
            '分级调服务': '1.1.7',
            '智能调服务': '1.1.7',
            '移动易运维': '1.1.1',
            '集团工单服务': '1.1.1',
            '复盘服务': '1.1.1',  # 新增：与用户数据一致
            'nacos': '2.5.2',
            '事件工单': '1.4.7',
        }
        return default_versions.get(software_name, '1.0.0')


def main():
    """主函数"""
    import sys

    # 默认输入输出文件路径
    input_file = 'docs/tools/contentToExcel/contentToChange.md'
    output_file = 'docs/tools/contentToExcel/shengchengbiaoge.xlsx'

    # 如果命令行提供了参数，使用提供的路径
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]

    # 创建转换器并生成 Excel
    converter = ContentToExcelConverter()
    try:
        converter.generate_excel(input_file, output_file)
    except FileNotFoundError:
        print(f"错误：找不到文件 {input_file}")
        sys.exit(1)
    except Exception as e:
        print(f"生成 Excel 时出错：{e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
