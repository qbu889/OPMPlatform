#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
从 Excel 导入调整因子数据到数据库
"""
import mysql.connector
from openpyxl import load_workbook
import os
from pathlib import Path
from dotenv import load_dotenv

# 加载 .env 文件
env_path = Path(__file__).parent.parent / '.env'
if env_path.exists():
    load_dotenv(env_path)
    print(f"✓ 已加载 .env 文件：{env_path}")

# 数据库配置
DB_CONFIG = {
    'host': os.getenv('MYSQL_HOST', 'localhost'),
    'port': int(os.getenv('MYSQL_PORT', 3306)),
    'user': os.getenv('MYSQL_USER', 'root'),
    'password': os.getenv('MYSQL_PASSWORD', '12345678'),
    'database': os.getenv('KNOWLEDGE_BASE_DB', 'knowledge_base'),
    'charset': os.getenv('MYSQL_CHARSET', 'utf8mb4')
}

def read_excel_data(file_path):
    """读取 Excel 文件数据"""
    wb = load_workbook(file_path)
    ws = wb.active
    
    data = []
    
    # 跳过标题行，从第 3 行开始（索引从 1 开始）
    for row_idx in range(3, ws.max_row + 1):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data.append(cell.value if cell.value else '')
        
        # 如果行非空，添加到数据列表
        if any(row_data):
            data.append(row_data)
    
    return data

def parse_adjustment_factors(data):
    """解析调整因子数据"""
    factors = []
    
    # 应用类型 (行 3, 12-19)
    # 因子名称 | 因子分类 | 选项名称 | 分值 | 公式
    application_types = [
        ('应用类型', '应用类型', '业务处理', 1.0, '=IF(C3=B12,D12,IF(C3=B13,D13,IF(C3=B14,D14,IF(C3=B15,D15,IF(C3=B16,D16,IF(C3=B17,D17,IF(C3=B18,D18,IF(C3=B19,D19,0))))))))'),
        ('应用类型', '应用类型', '应用集成', 1.2, '=IF(C3=B12,D12,IF(C3=B13,D13,IF(C3=B14,D14,IF(C3=B15,D15,IF(C3=B16,D16,IF(C3=B17,D17,IF(C3=B18,D18,IF(C3=B19,D19,0))))))))'),
        ('应用类型', '应用类型', '科技', 1.2, '=IF(C3=B12,D12,IF(C3=B13,D13,IF(C3=B14,D14,IF(C3=B15,D15,IF(C3=B16,D16,IF(C3=B17,D17,IF(C3=B18,D18,IF(C3=B19,D19,0))))))))'),
        ('应用类型', '应用类型', '多媒体', 1.5, '=IF(C3=B12,D12,IF(C3=B13,D13,IF(C3=B14,D14,IF(C3=B15,D15,IF(C3=B16,D16,IF(C3=B17,D17,IF(C3=B18,D18,IF(C3=B19,D19,0))))))))'),
        ('应用类型', '应用类型', '智能信息', 1.7, '=IF(C3=B12,D12,IF(C3=B13,D13,IF(C3=B14,D14,IF(C3=B15,D15,IF(C3=B16,D16,IF(C3=B17,D17,IF(C3=B18,D18,IF(C3=B19,D19,0))))))))'),
        ('应用类型', '应用类型', '基础软件/支撑软件', 1.7, '=IF(C3=B12,D12,IF(C3=B13,D13,IF(C3=B14,D14,IF(C3=B15,D15,IF(C3=B16,D16,IF(C3=B17,D17,IF(C3=B18,D18,IF(C3=B19,D19,0))))))))'),
        ('应用类型', '应用类型', '通信控制', 1.9, '=IF(C3=B12,D12,IF(C3=B13,D13,IF(C3=B14,D14,IF(C3=B15,D15,IF(C3=B16,D16,IF(C3=B17,D17,IF(C3=B18,D18,IF(C3=B19,D19,0))))))))'),
        ('应用类型', '应用类型', '流程控制', 2.0, '=IF(C3=B12,D12,IF(C3=B13,D13,IF(C3=B14,D14,IF(C3=B15,D15,IF(C3=B16,D16,IF(C3=B17,D17,IF(C3=B18,D18,IF(C3=B19,D19,0))))))))'),
    ]
    factors.extend(application_types)
    
    # 质量特性 - 分布式处理
    quality_factors = [
        ('分布式处理', '质量特性', '没有明示对分布式处理的需求事项', -1.0, '=IF(C4=C22,D22,IF(C4=C23,D23,IF(C4=C24,D24,-10)))'),
        ('分布式处理', '质量特性', '通过网络进行客户端/服务器及网络基础应用分布处理和传输', 0.0, '=IF(C4=C22,D22,IF(C4=C23,D23,IF(C4=C24,D24,-10)))'),
        ('分布式处理', '质量特性', '在多个服务器及处理器上同时相互执行计算机系统中的处理功能', 1.0, '=IF(C4=C22,D22,IF(C4=C23,D23,IF(C4=C24,D24,-10)))'),
        
        # 性能
        ('性能', '质量特性', '没有明示对性能的特别需求事项或活动，因此提供基本性能', -1.0, '=IF(C5=C25,D25,IF(C5=C26,D26,IF(C5=C27,D27,-10)))'),
        ('性能', '质量特性', '应答时间或处理率对高峰时间或所有业务时间来说都很重要，对连动系统结束处理时间的限制', 0.0, '=IF(C5=C25,D25,IF(C5=C26,D26,IF(C5=C27,D27,-10)))'),
        ('性能', '质量特性', '为满足性能需求事项，要求设计阶段开始进行性能分析，或在设计、开发阶段使用分析工具', 1.0, '=IF(C5=C25,D25,IF(C5=C26,D26,IF(C5=C27,D27,-10)))'),
        
        # 可靠性
        ('可靠性', '质量特性', '没有明示对可靠性的特别需求事项或活动，因此提供基本的可靠性', -1.0, '=IF(C6=C28,D28,IF(C6=C29,D29,IF(C6=C30,D30,-10)))'),
        ('可靠性', '质量特性', '发生故障时可轻易修复，带来一定不便或经济损失', 0.0, '=IF(C6=C28,D28,IF(C6=C29,D29,IF(C6=C30,D30,-10)))'),
        ('可靠性', '质量特性', '发生故障时很难复，发生重大经济损失或有生命危害', 1.0, '=IF(C6=C28,D28,IF(C6=C29,D29,IF(C6=C30,D30,-10)))'),
        
        # 多重站点
        ('多重站点', '质量特性', '在相同用途的硬件或软件环境下运行', -1.0, '=IF(C7=C31,D31,IF(C7=C32,D32,IF(C7=C33,D33,-10)))'),
        ('多重站点', '质量特性', '在用途类似的硬件或软件环境下运行', 0.0, '=IF(C7=C31,D31,IF(C7=C32,D32,IF(C7=C33,D33,-10)))'),
        ('多重站点', '质量特性', '在不同用途的硬件或软件环境下运行', 1.0, '=IF(C7=C31,D31,IF(C7=C32,D32,IF(C7=C33,D33,-10)))'),
    ]
    factors.extend(quality_factors)
    
    # 开发语言
    language_factors = [
        ('开发语言', '开发语言', 'C 及其他同级别语言/平台', 1.2, '=IF(C8=B43,D43,IF(C8=B44,D44,IF(C8=B45,D45,0)))'),
        ('开发语言', '开发语言', 'JAVA、C++、C#及其他同级别语言/平台', 1.0, '=IF(C8=B43,D43,IF(C8=B44,D44,IF(C8=B45,D45,0)))'),
        ('开发语言', '开发语言', 'PowerBuilder、ASP 及其他同级别语言/平台', 0.8, '=IF(C8=B43,D43,IF(C8=B44,D44,IF(C8=B45,D45,0)))'),
    ]
    factors.extend(language_factors)
    
    # 开发团队背景
    team_factors = [
        ('开发团队背景', '开发团队背景', '为本行业（政府）开发过类似的软件', 0.8, '=IF(C9=B48,D48,IF(C9=B49,D49,IF(C9=B50,D50,0)))'),
        ('开发团队背景', '开发团队背景', '为其他行业开发过类似的软件，或为本行业（政府）开发过不同但相关的软件', 1.0, '=IF(C9=B48,D48,IF(C9=B49,D49,IF(C9=B50,D50,0)))'),
        ('开发团队背景', '开发团队背景', '没有同类软件及本行业（政府）相关软件开发背景', 1.2, '=IF(C9=B48,D48,IF(C9=B49,D49,IF(C9=B50,D50,0)))'),
    ]
    factors.extend(team_factors)
    
    # 规模计数时机
    scale_factors = [
        ('规模计数时机', '规模变更调整系数', '估算早期', 1.39, '规模变更调整系数'),
        ('规模计数时机', '规模变更调整系数', '估算中期', 1.21, '规模变更调整系数'),
        ('规模计数时机', '规模变更调整系数', '估算晚期', 1.10, '规模变更调整系数'),
        ('规模计数时机', '规模变更调整系数', '项目完成', 1.00, '规模变更调整系数'),
    ]
    factors.extend(scale_factors)
    
    # 重用程度
    reuse_factors = [
        ('重用程度', '重用程度调整系数', '高', 0.3333, '=1/3'),
        ('重用程度', '重用程度调整系数', '中', 0.6667, '=2/3'),
        ('重用程度', '重用程度调整系数', '低', 1.0000, '=1/1'),
    ]
    factors.extend(reuse_factors)
    
    # 修改类型
    change_factors = [
        ('修改类型', '修改类型调整系数', '新增', 1.0000, '=1/1'),
        ('修改类型', '修改类型调整系数', '修改', 0.8000, '0.8'),
        ('修改类型', '修改类型调整系数', '删除', 0.2000, '0.2'),
    ]
    factors.extend(change_factors)
    
    return factors

def insert_to_database(factors):
    """插入数据到数据库"""
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        
        # 清空现有数据
        cursor.execute('TRUNCATE TABLE fpa_adjustment_factor')
        
        # 插入新数据
        sql = '''
            INSERT INTO fpa_adjustment_factor 
            (factor_name, factor_category, option_name, score_value, formula, display_order)
            VALUES (%s, %s, %s, %s, %s, %s)
        '''
        
        for idx, factor in enumerate(factors, 1):
            # factor 元组包含 5 个元素，加上 idx 作为 display_order
            cursor.execute(sql, factor + (idx,))
        
        conn.commit()
        
        # 验证插入的数据
        cursor.execute('SELECT COUNT(*) FROM fpa_adjustment_factor')
        count = cursor.fetchone()[0]
        
        print(f"✓ 成功插入 {count} 条调整因子数据")
        
        # 显示分类统计
        cursor.execute('''
            SELECT factor_category, COUNT(*) as count
            FROM fpa_adjustment_factor
            GROUP BY factor_category
            ORDER BY count DESC
        ''')
        
        print("\n📊 分类统计:")
        for row in cursor.fetchall():
            print(f"  - {row[0]}: {row[1]} 条")
        
        cursor.close()
        conn.close()
        
        return True
        
    except Exception as e:
        print(f"❌ 插入数据失败：{e}")
        return False

def main():
    """主函数"""
    print("=" * 60)
    print("调整因子数据导入工具")
    print("=" * 60)
    
    excel_path = 'test/fpa/调整因子.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"❌ Excel 文件不存在：{excel_path}")
        return
    
    print(f"\n📖 读取 Excel 文件：{excel_path}")
    data = read_excel_data(excel_path)
    print(f"✓ 读取到 {len(data)} 行数据")
    
    print("\n🔄 解析调整因子数据...")
    factors = parse_adjustment_factors(data)
    print(f"✓ 解析到 {len(factors)} 个调整因子")
    
    print("\n💾 插入数据到数据库...")
    if insert_to_database(factors):
        print("\n✅ 数据导入完成！")
    else:
        print("\n❌ 数据导入失败！")

if __name__ == '__main__':
    main()
