#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
内容转 Excel 工具 - 核心逻辑模块

功能：
1. 解析 .md/.txt 升级申请文档，提取多条升级记录
2. 根据日期规则匹配软件名称（3月→智能调服务，4月+→省内工单+智能调）
3. 按照公式链（N→L→O→P）计算标准版本号
4. 生成包含 4 个 Sheet 的 .xlsx 文件（含 Excel 公式）

Excel 公式说明：
- C列（版本号）=INT(O列)&"."&INT(MOD(O列,1)*10)&"."&MOD(INT(MOD(O列,1)*10),10)
- O列（版本计算）=IFERROR(N列+L列*0.01,0)
- L列（迭代次数）=COUNTIF($B$2:B列,B列)-1
- M列（初始版本）=IFERROR(VLOOKUP(B列,初始版本号!$A:$B,2,0),"未匹配软件")
- N列（计算小数）=IFERROR(LEFT(M列,FIND(".",M列)-1)+MID(M列,FIND(".",M列)+1,FIND(".",M列,FIND(".",M列)+1)-FIND(".",M列)-1)/10+RIGHT(M列,LEN(M列)-FIND(".",M列,FIND(".",M列)+1))/100,0)
- P列（标准版本号）=IF(O列=0,"无版本",INT(O列)&"."&INT(MOD(O列,1)*10)&"."&MOD(INT(MOD(O列,1)*10),10))

作者: Claude Code
日期: 2026-05-29
"""

import re
import os
import uuid
from datetime import datetime
from typing import Dict, List, Optional

from openpyxl import Workbook


# ============================================================================
# 内置软件名称关键字库（13条完整记录，与用户提供的数据一致）
# ============================================================================

BUILTIN_SOFTWARE_MAP: Dict[str, Dict] = {
    '省内工单服务': {
        'keywords': ['省内工单', '工单服务'],
        'target_ip': '10.44.225.197、10.43.118.48、10.43.118.47、10.44.225.30、10.44.225.31、10.44.225.33',
        'operator': '姚翔',
        'verifier': '林子旺',
    },
    '告警数据采集': {
        'keywords': ['告警数据', '告警采集'],
        'target_ip': '10.43.118.45、10.43.118.46、10.44.225.116',
        'operator': '高胜永',
        'verifier': '林子旺',
    },
    '重定义服务': {
        'keywords': ['重定义服务', '重定义'],
        'target_ip': '10.43.118.45、10.43.118.46、10.44.225.116',
        'operator': '高胜永',
        'verifier': '林子旺',
    },
    '产品重定义服务': {
        'keywords': ['产品重定义', '产品业务事件监控'],
        'target_ip': '10.43.118.45、10.43.118.46、10.44.225.116',
        'operator': '高胜永',
        'verifier': '林子旺',
    },
    '实例化 es 入库': {
        'keywords': ['es 入库', '实例化 es'],
        'target_ip': '10.43.118.45、10.43.118.46、10.44.225.116',
        'operator': '高胜永',
        'verifier': '林子旺',
    },
    '规则匹配服务': {
        'keywords': ['规则匹配'],
        'target_ip': '10.43.118.45、10.43.118.46、10.44.225.116',
        'operator': '高胜永',
        'verifier': '林子旺',
    },
    '事件查询服务': {
        'keywords': ['事件查询'],
        'target_ip': '10.43.118.45、10.43.118.46、10.44.225.116',
        'operator': '高胜永',
        'verifier': '林子旺',
    },
    '分级调服务': {
        'keywords': ['分级调'],
        'target_ip': '10.43.52.51、10.43.52.52、10.43.52.53、10.43.52.54',
        'operator': '高胜永',
        'verifier': '林子旺',
    },
    '智能调服务': {
        'keywords': ['智能调度', '智能调'],
        'target_ip': '10.46.102.85、10.46.102.86',
        'operator': '李金山',
        'verifier': '林子旺',
    },
    '移动易运维': {
        'keywords': ['移动易'],
        'target_ip': '10.43.52.58、10.43.52.59、10.43.149.199',
        'operator': '张保全',
        'verifier': '林子旺',
    },
    'nacos': {
        'keywords': ['nacos'],
        'target_ip': '10.43.52.58、10.43.52.59、10.43.149.199',
        'operator': '张保全',
        'verifier': '林子旺',
    },
    '集团工单服务': {
        'keywords': ['集团工单'],
        'target_ip': '10.44.225.32、10.44.225.39、10.46.102.90、10.46.102.91、10.46.102.92',
        'operator': '金文辉',
        'verifier': '林子旺',
    },
    '复盘服务': {
        'keywords': ['复盘'],
        'target_ip': '10.44.225.30、10.44.225.31、10.44.225.33、10.46.102.83、10.46.102.84',
        'operator': '金文辉',
        'verifier': '林子旺',
    },
}

BUILTIN_INITIAL_VERSIONS: Dict[str, str] = {
    '省内工单服务': '1.4.6',
    '告警数据采集': '1.1.5',
    '重定义服务': '1.1.5',
    '产品重定义服务': '1.1.5',
    '实例化 es 入库': '1.1.5',
    '规则匹配服务': '1.1.5',
    '事件查询服务': '1.1.5',
    '分级调服务': '1.1.7',
    '智能调服务': '1.1.7',
    '移动易运维': '1.1.1',
    'nacos': '2.5.2',
    '集团工单服务': '1.1.1',
    '复盘服务': '1.1.0',
}


# ============================================================================
# 升级申请标题正则
# ============================================================================

UPGRADE_TITLE_PATTERN = re.compile(
    r'监控综合应用平台(\d{4})年(\d{1,2})月(\d{1,2})日(功能升级|服务迁移|安全加固)申请'
)

# 日期修正规则：3月2日实际对应3月11日（历史修正）
DATE_CORRECTIONS = {
    (2026, 3, 2): (2026, 3, 11),
}


# ============================================================================
# 核心解析函数（供 routes 调用）
# ============================================================================

def parse_upgrade_records(content: str) -> List[Dict]:
    """
    解析 markdown/txt 内容，提取所有升级申请记录。

    Args:
        content: 文件内容字符串

    Returns:
        升级记录列表，每条包含 year, month, day, date_str, type_desc, title
    """
    records = []
    matches = list(UPGRADE_TITLE_PATTERN.finditer(content))

    for i, match in enumerate(matches):
        year = int(match.group(1))
        month = int(match.group(2))
        day = int(match.group(3))

        # 应用日期修正规则：3月2日→3月11日
        correction_key = (year, month, day)
        if correction_key in DATE_CORRECTIONS:
            year, month, day = DATE_CORRECTIONS[correction_key]

        type_desc = match.group(4)
        date_str = f'{year}/{month}/{day}'

        # 提取该记录之后的内容（用于软件匹配）
        next_start = matches[i + 1].start() if i + 1 < len(matches) else len(content)
        section_content = content[match.end():next_start]

        records.append({
            'year': year,
            'month': month,
            'day': day,
            'date_str': date_str,
            'type_desc': type_desc,
            'title': f'监控综合应用平台{year}年{month}月{day}日{type_desc}申请',
            'section_content': section_content,
        })

    return records


def match_softwares(record: Dict, software_ip_map: Optional[Dict] = None) -> List[str]:
    """
    根据升级记录的日期和内容匹配涉及的软件。

    匹配策略：
    1. 关键字匹配：从升级内容中查找软件关键字（优先）
    2. 日期规则回退：如果关键字匹配为空，则使用日期规则（3月→智能调服务，4月+→省内工单+智能调）

    Args:
        record: 升级记录字典
        software_ip_map: 外部软件-IP映射（优先使用），None则回退到内置

    Returns:
        匹配到的软件名称列表
    """
    month = record['month']
    section_content = record.get('section_content', '')

    # 获取软件映射（优先外部，回退内置）
    if software_ip_map:
        sw_map = software_ip_map
    else:
        sw_map = BUILTIN_SOFTWARE_MAP

    # 关键字匹配：从内容中查找所有匹配的软件
    matched = set()
    for sw_name, info in sw_map.items():
        keywords = info.get('keywords', [])
        for kw in keywords:
            if kw.lower() in section_content.lower():
                matched.add(sw_name)

    # 如果关键字匹配到了结果，直接返回
    if matched:
        return sorted(matched)

    # 回退到日期规则（兼容旧版文档）
    if month == 3:
        return ['智能调服务']
    else:
        return ['省内工单服务', '智能调服务']


def calculate_standard_version(initial_version: str, iteration_count: int) -> str:
    """
    按照公式链计算标准版本号。

    公式链：
      N (迭代次数) = COUNTIF(该软件在表中出现次数) - 1
      L (计算小数) = 解析初始版本"主.次.修"为小数
      O (版本计算) = N + L * 0.01
      P (标准版本号) = INT(O).INT(MOD(O,1)*10).MOD(INT(MOD(O,1)*10),10)

    Args:
        initial_version: 初始版本号，如 "1.2.1"
        iteration_count: 迭代次数（从0开始）

    Returns:
        标准版本号字符串，如 "1.2.2"
    """
    parts = initial_version.split('.')
    if len(parts) < 3:
        return initial_version

    major = int(parts[0])
    minor = int(parts[1])
    patch = int(parts[2])

    # 计算小数 L = major + minor/10 + patch/100
    decimal_value = major + minor / 10.0 + patch / 100.0

    # 版本计算 O = iteration_count + decimal_value * 0.01
    version_calc = iteration_count + decimal_value * 0.01

    # 标准版本号 P
    if version_calc == 0:
        return '无版本'

    p_major = int(version_calc)
    p_minor = int(version_calc % 1 * 10)
    p_patch = int(int(version_calc % 1 * 10) % 10)

    return f'{p_major}.{p_minor}.{p_patch}'


def generate_excel(content: str, software_ip_map: Optional[Dict] = None) -> str:
    """
    根据升级申请内容生成包含 4 个 Sheet 的 Excel 文件。

    Args:
        content: 升级申请文档内容字符串
        software_ip_map: 外部软件-IP映射（可选）

    Returns:
        生成的 Excel 文件路径
    """
    # 解析升级记录
    records = parse_upgrade_records(content)

    if not records:
        raise ValueError('未解析到任何升级申请记录，请检查文件格式是否符合规范。')

    # 获取软件映射
    if software_ip_map:
        sw_map = software_ip_map
    else:
        sw_map = BUILTIN_SOFTWARE_MAP

    # 获取初始版本映射
    if software_ip_map:
        initial_versions = {
            name: info.get('initial_version', BUILTIN_INITIAL_VERSIONS.get(name, '1.0.0'))
            for name, info in sw_map.items()
        }
    else:
        initial_versions = BUILTIN_INITIAL_VERSIONS

    # 创建 Workbook
    wb = Workbook()

    # ========================================================================
    # Sheet 1: 主表（202603～5月）
    # ========================================================================
    ws_main = wb.active
    if ws_main:
        ws_main.title = '202603～5月'

    main_headers = [
        '序号', '软件名称', '版本号', '装载日期', '操作类型',
        '目标名称', '目标IP', '操作人员', '操作验证', '验证人员', '备注',
        '迭代次数 N', '首次初始版本', '计算小数', '版本计算', '标准版本号'
    ]
    ws_main.append(main_headers)

    # 跟踪每个软件的迭代次数
    software_iteration = {}
    row_counter = 0

    for record in records:
        month = record['month']
        date_str = record['date_str']
        install_date = f'{date_str} 22:00'

        # 根据日期匹配软件
        matched_softwares = match_softwares(record, sw_map)

        for software in matched_softwares:
            row_num = row_counter + 2  # Excel行号（第1行为表头）
            row_counter += 1

            # 获取软件信息
            sw_info = sw_map.get(software, {})
            target_ip = sw_info.get('target_ip', '')
            operator = sw_info.get('operator', '林子旺')
            verifier = sw_info.get('verifier', '林子旺')

            # 初始化迭代计数
            if software not in software_iteration:
                software_iteration[software] = 0

            iteration = software_iteration[software]
            initial_version = initial_versions.get(software, '1.0.0')

            # 计算标准版本号
            std_version = calculate_standard_version(initial_version, iteration)

            # A列：序号
            ws_main[f'A{row_num}'] = row_counter

            # B列：软件名称
            ws_main[f'B{row_num}'] = software

            # C列：版本号（公式）
            ws_main[f'C{row_num}'] = (
                f'=IF(O{row_num}=0,"无版本",INT(O{row_num})&"."&'
                f'INT(MOD(O{row_num},1)*10)&"."&MOD(INT(MOD(O{row_num},1)*10),10)'
            )

            # D列：装载日期
            ws_main[f'D{row_num}'] = install_date

            # E列：操作类型
            type_desc = record['type_desc']
            if type_desc in ('功能升级', '服务迁移'):
                op_type = '更新'
            elif type_desc == '安全加固':
                op_type = '安全加固'
            else:
                op_type = '更新'
            ws_main[f'E{row_num}'] = op_type

            # F列：目标名称（空值预留）
            ws_main[f'F{row_num}'] = ''

            # G列：目标IP（VLOOKUP公式）- 第3列（添加序号后偏移）
            ws_main[f'G{row_num}'] = f'=VLOOKUP(B{row_num},软件名称和目标IP!$B:$F,3,FALSE)'

            # H列：操作人员（VLOOKUP公式）- 第4列（添加序号后偏移）
            ws_main[f'H{row_num}'] = f'=VLOOKUP(B{row_num},软件名称和目标IP!$B:$F,4,FALSE)'

            # I列：操作验证（固定值）
            ws_main[f'I{row_num}'] = '正常'

            # J列：验证人员（VLOOKUP公式）
            ws_main[f'J{row_num}'] = f'=VLOOKUP(B{row_num},软件名称和目标IP!$B:$F,6,FALSE)'

            # K列：备注（申请标题拼接）
            ws_main[f'K{row_num}'] = (
                f'="监控综合应用平台"&YEAR(D{row_num})&"月"&DAY(D{row_num})&"日功能升级申请"'
            )

            # L列：迭代次数 N（公式）
            ws_main[f'L{row_num}'] = f'=COUNTIF($B$2:B{row_num},B{row_num})-1'

            # M列：首次初始版本（VLOOKUP公式）
            ws_main[f'M{row_num}'] = (
                f'=IFERROR(VLOOKUP(B{row_num},初始版本号!$A:$B,2,0),"未匹配软件")'
            )

            # N列：计算小数（公式）
            ws_main[f'N{row_num}'] = (
                f'=IFERROR(LEFT(M{row_num},FIND(".",M{row_num})-1)+'
                f'MID(M{row_num},FIND(".",M{row_num})+1,'
                f'FIND(".",M{row_num},FIND(".",M{row_num})+1)-FIND(".",M{row_num})-1)/10+'
                f'RIGHT(M{row_num},LEN(M{row_num})-FIND(".",M{row_num},FIND(".",M{row_num})+1))/100,0)'
            )

            # O列：版本计算（公式）
            ws_main[f'O{row_num}'] = f'=IFERROR(N{row_num}+L{row_num}*0.01,0)'

            # P列：标准版本号（公式）
            ws_main[f'P{row_num}'] = (
                f'=IF(O{row_num}=0,"无版本",INT(O{row_num})&"."&'
                f'INT(MOD(O{row_num},1)*10)&"."&MOD(INT(MOD(O{row_num},1)*10),10))'
            )

            software_iteration[software] += 1

    # 设置列宽
    col_widths = {
        'A': 8, 'B': 15, 'C': 12, 'D': 18, 'E': 10,
        'F': 10, 'G': 30, 'H': 10, 'I': 10, 'J': 25,
        'K': 35, 'L': 10, 'M': 12, 'N': 12, 'O': 12, 'P': 12
    }
    for col_letter, width in col_widths.items():
        ws_main.column_dimensions[col_letter].width = width

    # ========================================================================
    # Sheet 2: 版本号计算（含历史+新增）
    # ========================================================================
    ws_version = wb.create_sheet('版本号计算')

    version_headers = [
        '序号', '软件名称', '版本号', '装载日期', '操作类型',
        '目标IP', '操作人员', '操作验证', '验证人员', '备注',
        '迭代次数 N', '首次初始版本', '计算小数', '版本计算', '标准版本号'
    ]
    ws_version.append(version_headers)

    version_iteration = {}
    v_row_counter = 0

    for record in records:
        month = record['month']
        date_str = record['date_str']
        install_date = f'{date_str} 22:00'

        matched_softwares = match_softwares(record, sw_map)

        for software in matched_softwares:
            row_num = v_row_counter + 2
            v_row_counter += 1

            if software not in version_iteration:
                version_iteration[software] = 0

            iteration = version_iteration[software]
            initial_version = initial_versions.get(software, '1.0.0')
            std_version = calculate_standard_version(initial_version, iteration)

            sw_info = sw_map.get(software, {})
            operator = sw_info.get('operator', '林子旺')
            verifier = sw_info.get('verifier', '林子旺')

            ws_version[f'A{row_num}'] = v_row_counter
            ws_version[f'B{row_num}'] = software
            ws_version[f'C{row_num}'] = std_version
            ws_version[f'D{row_num}'] = install_date

            type_desc = record['type_desc']
            if type_desc in ('功能升级', '服务迁移'):
                op_type = '更新'
            elif type_desc == '安全加固':
                op_type = '安全加固'
            else:
                op_type = '更新'
            ws_version[f'E{row_num}'] = op_type

            ws_version[f'F{row_num}'] = sw_info.get('target_ip', '')
            # 操作人员（VLOOKUP公式）- 第4列（添加序号后偏移）
            ws_version[f'G{row_num}'] = f'=VLOOKUP(B{row_num},软件名称和目标IP!$B:$F,4,FALSE)'
            ws_version[f'H{row_num}'] = '正常'
            ws_version[f'I{row_num}'] = verifier

            ws_version[f'J{row_num}'] = (
                f'="监控综合应用平台"&YEAR(D{row_num})&"月"&DAY(D{row_num})&"日"&E{row_num}&"申请"'
            )

            ws_version[f'K{row_num}'] = f'=COUNTIF($B$2:B{row_num},B{row_num})-1'
            ws_version[f'L{row_num}'] = (
                f'=IFERROR(VLOOKUP(B{row_num},初始版本号!$A:$B,2,0),"未匹配软件")'
            )

            ws_version[f'M{row_num}'] = (
                f'=IFERROR(LEFT(L{row_num},FIND(".",L{row_num})-1)+'
                f'MID(L{row_num},FIND(".",L{row_num})+1,'
                f'FIND(".",L{row_num},FIND(".",L{row_num})+1)-FIND(".",L{row_num})-1)/10+'
                f'RIGHT(L{row_num},LEN(L{row_num})-FIND(".",L{row_num},FIND(".",L{row_num})+1))/100,0)'
            )

            ws_version[f'N{row_num}'] = f'=IFERROR(M{row_num}+K{row_num}*0.01,0)'
            ws_version[f'O{row_num}'] = (
                f'=IF(N{row_num}=0,"无版本",INT(N{row_num})&"."&'
                f'INT(MOD(N{row_num},1)*10)&"."&MOD(INT(MOD(N{row_num},1)*10),10))'
            )

            version_iteration[software] += 1

    for col_letter, width in col_widths.items():
        ws_version.column_dimensions[col_letter].width = width

    # ========================================================================
    # Sheet 3: 初始版本号（基准参考表）
    # ========================================================================
    ws_initial = wb.create_sheet('初始版本号')
    ws_initial.append(['软件名称', '初始化基准版本'])

    for name in sorted(initial_versions.keys()):
        ws_initial.append([name, initial_versions[name]])

    ws_initial.column_dimensions['A'].width = 20
    ws_initial.column_dimensions['B'].width = 15

    # ========================================================================
    # Sheet 4: 软件名称和目标IP（映射参考表）
    # ========================================================================
    ws_ip = wb.create_sheet('软件名称和目标IP')
    ws_ip.append(['序号', '软件名称', '目标IP', '操作人员', '验证人员', '版本号'])

    for idx, (name, info) in enumerate(sorted(sw_map.items()), 1):
        initial_ver = initial_versions.get(name, '')
        ws_ip.append([
            idx, name, info.get('target_ip', ''),
            info.get('operator', ''), info.get('verifier', ''), initial_ver
        ])

    ws_ip.column_dimensions['A'].width = 8
    ws_ip.column_dimensions['B'].width = 15
    ws_ip.column_dimensions['C'].width = 90
    ws_ip.column_dimensions['D'].width = 10
    ws_ip.column_dimensions['E'].width = 10
    ws_ip.column_dimensions['F'].width = 12

    # ========================================================================
    # 保存文件
    # ========================================================================
    output_dir = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), '..', 'uploads', 'content_to_excel'
    )
    os.makedirs(output_dir, exist_ok=True)

    file_name = f'content_to_excel_{uuid.uuid4().hex[:8]}.xlsx'
    output_path = os.path.join(output_dir, file_name)

    wb.save(output_path)
    return output_path


# ============================================================================
# 兼容旧版接口（供直接脚本调用）
# ============================================================================

def parse_content_to_change(file_path: str) -> List[Dict]:
    """解析 contentToChange.md 文件，提取升级信息（兼容旧接口）"""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    return parse_upgrade_records(content)


def calculate_version(initial_version: str, iteration_count: int) -> str:
    """计算版本号（兼容旧接口）"""
    parts = initial_version.split('.')
    if len(parts) == 3:
        major, minor, patch = int(parts[0]), int(parts[1]), int(parts[2])
        new_patch = patch + iteration_count

        if new_patch >= 10:
            minor += new_patch // 10
            new_patch = new_patch % 10

        if minor >= 10:
            major += minor // 10
            minor = minor % 10

        return f'{major}.{minor}.{new_patch}'
    return initial_version


def generate_excel_from_file(input_file: str, output_file: str):
    """从文件生成 Excel（兼容旧接口）"""
    with open(input_file, 'r', encoding='utf-8') as f:
        content = f.read()

    records = parse_upgrade_records(content)

    wb = Workbook()
    ws_main = wb.active
    if ws_main:
        ws_main.title = '202603～5月'

    headers = [
        '序号', '软件名称', '版本号', '装载日期', '操作类型', '目标名称',
        '目标IP', '操作人员', '操作验证', '验证人员', '备注'
    ]
    ws_main.append(headers)

    software_iterations = {name: 0 for name in BUILTIN_SOFTWARE_MAP}
    row_idx = 1

    for upgrade in records:
        if upgrade['month'] == 3:
            softwares = ['智能调服务']
        else:
            softwares = ['省内工单服务', '智能调服务']

        for software in softwares:
            info = BUILTIN_SOFTWARE_MAP[software]
            iteration = software_iterations.get(software, 0)

            parts = info['initial_version'].split('.')
            if len(parts) == 3:
                major, minor, patch = int(parts[0]), int(parts[1]), int(parts[2])
                new_patch = patch + iteration + 1

                if new_patch >= 10:
                    minor += new_patch // 10
                    new_patch = new_patch % 10

                if minor >= 10:
                    major += minor // 10
                    minor = minor % 10

                version = f'{major}.{minor}.{new_patch}'
            else:
                version = info['initial_version']

            ws_main.append([
                row_idx, software, version,
                f"{upgrade['year']}/{upgrade['month']}/{upgrade['day']} 22:00",
                '更新', '', info['target_ip'],
                info['operator'], '正常', info['verifier'],
                upgrade['title']
            ])

            software_iterations[software] += 1
            row_idx += 1

    ws_main.column_dimensions['A'].width = 8
    ws_main.column_dimensions['B'].width = 15
    ws_main.column_dimensions['C'].width = 10
    ws_main.column_dimensions['D'].width = 20
    ws_main.column_dimensions['E'].width = 10
    ws_main.column_dimensions['F'].width = 10
    ws_main.column_dimensions['G'].width = 90
    ws_main.column_dimensions['H'].width = 10
    ws_main.column_dimensions['I'].width = 10
    ws_main.column_dimensions['J'].width = 10
    ws_main.column_dimensions['K'].width = 40

    # 版本号计算 sheet
    ws_version = wb.create_sheet('版本号计算')
    ws_version.append(['软件名称', '初始版本', '升级次数', '计算过程', '最终版本'])

    software_iterations2 = {name: 0 for name in BUILTIN_SOFTWARE_MAP}
    for upgrade in records:
        if upgrade['month'] == 3:
            softwares = ['智能调服务']
        else:
            softwares = ['省内工单服务', '智能调服务']

        for software in softwares:
            info = BUILTIN_SOFTWARE_MAP[software]
            iteration = software_iterations2.get(software, 0)

            parts = info['initial_version'].split('.')
            if len(parts) == 3:
                major, minor, patch = int(parts[0]), int(parts[1]), int(parts[2])
                new_patch = patch + iteration + 1

                if new_patch >= 10:
                    minor += new_patch // 10
                    new_patch = new_patch % 10

                if minor >= 10:
                    major += minor // 10
                    minor = minor % 10

                version = f'{major}.{minor}.{new_patch}'
            else:
                version = info['initial_version']

            ws_version.append([
                software, info['initial_version'], iteration + 1,
                f"{info['initial_version']} + {iteration + 1}次升级", version
            ])

            software_iterations2[software] += 1

    ws_version.column_dimensions['A'].width = 15
    ws_version.column_dimensions['B'].width = 12
    ws_version.column_dimensions['C'].width = 10
    ws_version.column_dimensions['D'].width = 30
    ws_version.column_dimensions['E'].width = 12

    # 初始版本号 sheet
    ws_initial = wb.create_sheet('初始版本号')
    ws_initial.append(['软件名称', '初始化基准版本'])
    for software, info in BUILTIN_SOFTWARE_MAP.items():
        ws_initial.append([software, info['initial_version']])

    # 软件名称和目标IP sheet（添加序号列）
    ws_ip = wb.create_sheet('软件名称和目标IP')
    ws_ip.append(['序号', '软件名称', '目标IP', '操作人员', '验证人员'])
    for idx, (software, info) in enumerate(BUILTIN_SOFTWARE_MAP.items(), 1):
        ws_ip.append([idx, software, info['target_ip'], info['operator'], info['verifier']])

    wb.save(output_file)
    print(f'Excel文件已生成: {output_file}')


if __name__ == '__main__':
    import sys

    base_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.join(base_dir, '..', '..')

    input_path = os.path.join(
        project_root, 'docs', 'tools', 'contentToExcel', 'contentToChange.md'
    )
    output_path = os.path.join(
        project_root, 'docs', 'tools', 'contentToExcel', 'shengchengbiaoge.xlsx'
    )

    if len(sys.argv) >= 2:
        input_path = sys.argv[1]
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]

    generate_excel_from_file(input_path, output_path)
