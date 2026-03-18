#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
调整因子管理路由
"""
from flask import Blueprint, render_template, request, jsonify, current_app
import mysql.connector
from decimal import Decimal
import json
from datetime import datetime

adjustment_bp = Blueprint('adjustment', __name__, url_prefix='/adjustment')


def get_db_connection():
    """获取数据库连接"""
    return mysql.connector.connect(
        host=current_app.config['MYSQL_HOST'],
        port=current_app.config['MYSQL_PORT'],
        user=current_app.config['MYSQL_USER'],
        password=current_app.config['MYSQL_PASSWORD'],
        charset=current_app.config['MYSQL_CHARSET'],
        database=current_app.config['KNOWLEDGE_BASE_DB']
    )


def decimal_to_float(obj):
    """将 Decimal 转换为 float"""
    if isinstance(obj, Decimal):
        return float(obj)
    return obj


@adjustment_bp.route('/page')
def adjustment_factor_page():
    """调整因子管理页面"""
    return render_template('adjustment_factor.html')


@adjustment_bp.route('/api/factors', methods=['GET'])
def get_factors():
    """获取所有调整因子"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute('''
            SELECT * FROM fpa_adjustment_factor 
            ORDER BY factor_category, display_order, id
        ''')
        
        factors = cursor.fetchall()
        
        # 转换 Decimal 为 float
        for factor in factors:
            if factor.get('score_value'):
                factor['score_value'] = decimal_to_float(factor['score_value'])
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'data': factors
        })
        
    except Exception as e:
        current_app.logger.error(f"获取调整因子失败：{e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@adjustment_bp.route('/api/factor', methods=['POST'])
def create_factor():
    """创建调整因子"""
    try:
        data = request.json
        factor_name = data.get('factor_name')
        factor_category = data.get('factor_category')
        option_name = data.get('option_name', '')
        score_value = data.get('score_value', 0)
        formula = data.get('formula', '')
        display_order= data.get('display_order', 0)
        parent_id = data.get('parent_id')
        
        if not factor_name:
            return jsonify({
                'success': False,
                'message': '因子名称不能为空'
            }), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO fpa_adjustment_factor 
            (factor_name, factor_category, option_name, score_value, formula, display_order, parent_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        ''', (factor_name, factor_category, option_name, score_value, formula, display_order, parent_id))
        
        factor_id = cursor.lastrowid
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '添加成功',
            'data': {'id': factor_id}
        })
        
    except Exception as e:
        current_app.logger.error(f"创建调整因子失败：{e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@adjustment_bp.route('/api/factor/<int:factor_id>', methods=['PUT'])
def update_factor(factor_id):
    """更新调整因子"""
    try:
        data = request.json
        factor_name = data.get('factor_name')
        factor_category = data.get('factor_category')
        option_name = data.get('option_name', '')
        score_value = data.get('score_value', 0)
        formula = data.get('formula', '')
        display_order = data.get('display_order', 0)
        parent_id = data.get('parent_id')
        
        if not factor_name:
            return jsonify({
                'success': False,
                'message': '因子名称不能为空'
            }), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE fpa_adjustment_factor 
            SET factor_name = %s,
                factor_category = %s,
                option_name = %s,
                score_value = %s,
                formula = %s,
                display_order = %s,
                parent_id = %s
            WHERE id = %s
        ''', (factor_name, factor_category, option_name, score_value, formula, 
              display_order, parent_id, factor_id))
        
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '更新成功'
        })
        
    except Exception as e:
        current_app.logger.error(f"更新调整因子失败：{e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@adjustment_bp.route('/api/factor/<int:factor_id>', methods=['DELETE'])
def delete_factor(factor_id):
    """删除调整因子"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 先删除子项
        cursor.execute('DELETE FROM fpa_adjustment_factor WHERE parent_id = %s', (factor_id,))
        # 删除主项
        cursor.execute('DELETE FROM fpa_adjustment_factor WHERE id = %s', (factor_id,))
        
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '删除成功'
        })
        
    except Exception as e:
        current_app.logger.error(f"删除调整因子失败：{e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@adjustment_bp.route('/api/config', methods=['GET'])
def get_config():
    """获取调整因子配置"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute('SELECT * FROM fpa_adjustment_config ORDER BY config_type, config_key')
        configs = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'data': configs
        })
        
    except Exception as e:
        current_app.logger.error(f"获取配置失败：{e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@adjustment_bp.route('/api/config', methods=['POST'])
def save_config():
    """保存调整因子配置"""
    try:
        data = request.json
        config_key = data.get('config_key')
        config_value = data.get('config_value')
        config_type = data.get('config_type')
        description = data.get('description', '')
        
        if not config_key or not config_value:
            return jsonify({
                'success': False,
                'message': '配置键和配置值不能为空'
            }), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO fpa_adjustment_config 
            (config_key, config_value, config_type, description)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE 
                config_value = VALUES(config_value),
                config_type = VALUES(config_type),
                description = VALUES(description),
                updated_at = CURRENT_TIMESTAMP
        ''', (config_key, config_value, config_type, description))
        
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '保存成功'
        })
        
    except Exception as e:
        current_app.logger.error(f"保存配置失败：{e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@adjustment_bp.route('/api/import-excel', methods=['POST'])
def import_from_excel():
    """从 Excel 导入调整因子"""
    try:
        from openpyxl import load_workbook
        import os
        
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'message': '没有上传文件'
            }), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({
                'success': False,
                'message': '文件名为空'
            }), 400
        
        # 保存临时文件
        temp_path = os.path.join(current_app.config['TMP_FOLDER'], 'adjustment_factor_temp.xlsx')
        os.makedirs(os.path.dirname(temp_path), exist_ok=True)
        file.save(temp_path)
        
        # 读取 Excel
        wb = load_workbook(temp_path)
        ws = wb.active
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 清空现有数据
        cursor.execute('TRUNCATE TABLE fpa_adjustment_factor')
        
        # 解析 Excel 数据（从第 3 行开始，跳过标题）
        row_num = 3
        while row_num <= ws.max_row:
            row = [cell.value for cell in ws.row_dimensions[row_num]]
            
            # 如果行为空，跳过
            if not any(row):
                row_num += 1
                continue
            
            factor_name = row[0] if len(row) > 0 else None
            factor_category = row[1] if len(row) > 1 else None
            option_name = row[2] if len(row) > 2 else None
            formula = row[3] if len(row) > 3 else None
            
            # 提取分值（从公式或固定值）
            score_value = 0
            if formula and 'D' in formula:
                # 尝试从公式中提取引用单元格
                pass
            
            if factor_name:
                cursor.execute('''
                    INSERT INTO fpa_adjustment_factor 
                    (factor_name, factor_category, option_name, score_value, formula, display_order)
                    VALUES (%s, %s, %s, %s, %s, %s)
                ''', (factor_name, factor_category, option_name, score_value, formula, row_num))
            
            row_num += 1
        
        conn.commit()
        
        cursor.close()
        conn.close()
        
        # 删除临时文件
        os.remove(temp_path)
        
        return jsonify({
            'success': True,
            'message': f'成功导入 {row_num - 3} 条数据'
        })
        
    except Exception as e:
        current_app.logger.error(f"从 Excel 导入失败：{e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500
