# routes/fpa_generator_routes.py
from flask import Blueprint, render_template, request, jsonify, send_file, flash, current_app
from pathlib import Path
import pandas as pd
import re
import os
import time
from datetime import datetime
import logging
import mysql.connector
from decimal import Decimal
from .fpa_ai_expander import ai_assisted_expand_function_points
from models.fpa_category_rules import FPACategoryRule

fpa_generator_bp = Blueprint('fpa_generator', __name__, url_prefix='/fpa-generator')
logger = logging.getLogger(__name__)

# ========== 预编译正则表达式（性能优化） ==========
# 在模块加载时只编译一次，避免重复编译开销
PATTERNS = {
    'level1': re.compile(r'^##\s+(.+)$'),  # 一级分类
    'level2': re.compile(r'^###\s+(.+)$'),  # 二级分类
    'level3': re.compile(r'^####\s+(.+)$'),  # 三级分类
    'level4': re.compile(r'^#####\s+(.+)$'),  # 功能点名称
    'level5': re.compile(r'^######\s*(.+?)(?:（注.*?）)?$'),  # 功能点计数项
    'bold': re.compile(r'\*\*'),  # 加粗符号
    'note_zh': re.compile(r'[（ (]注.*?[)）]'),  # 中文注释
    'note_end': re.compile(r'\s*[（ (]注.*?[)）]\s*$'),  # 末尾注释
    'zero_width': re.compile(r'[\u200B\u200D]'),  # 零宽字符
    'special_chars': re.compile(r'[^\u4e00-\u9fa5a-zA-Z0-9_\.\s]'),  # 特殊字符
    'func_desc': re.compile(r'[：**]\s*(.+?)(?:[：**]|$)'),  # 字段值提取
    'ilf_count': re.compile(r'涉及到 (\d+) 个内部逻辑文件'),  # ILF 数量
    'elf_count': re.compile(r'，(\d+) 个外部逻辑文件'),  # ELF 数量
    'ilf_files': re.compile(r'本期新增/变更的内部逻辑文件\s*[:：]\s*(.+)'),  # ILF 文件列表
    'table_split': re.compile(r'[,,\u3001]'),  # 表格分隔符
    'category_note': re.compile(r'\s*[（ (]注.*?[)）]\s*$'),  # 类别注释
}

# 关键词匹配模式（用于类别识别）
KEYWORD_PATTERNS = {
    'ilf_suffix': re.compile(r'.*表$'),  # ILF 后缀
    'ilf_keywords': re.compile(r'数据表 | 配置表 | 结果表 | 详单表'),
    'ei_keywords': re.compile(r'录入 | 修改 | 删除 | 增 | 删 | 改 | 同步 | 导入 | 添加 | 设置 | 保存 | 提交 | 移交 | 回单 | 赋值'),
    'eo_keywords': re.compile(r'判定 | 分析 | 计算 | 处理 | 识别 | 匹配 | 切换 | 导出 | 上报 | 调度 | 推送 | 验证 | 检测 | 剔除 | 运算 | 渲染 | 生成 | 跳转 | 控制 | 监听 | 播报 | 触发 | 过滤 | 建议输出 | 排查 | 关联 | 复盘 | 审核 | 流转 | 总结 | 报告 | 标签输出 | 执行情况 | 存在问题 | 简要说明 | 照片上传 | 自动流转 | 消息通知 | 确认 | 驳回 | 归档 | 选择 | 执行 | 映射 | 采集 | 自动派发 | 人工派发 | 配置化 | 下钻'),
    'eq_keywords': re.compile(r'列表 | 快速查询 | 查询 | 搜索 | 查看 | 浏览 | 筛选 | 详情 | 展示 | 显示 | 获取 | 读取'),
    'present_keywords': re.compile(r'呈现'),
    'present_eq_keywords': re.compile(r'关联 | 隐患 | 规则 | 列表'),
}
# ================================================

# FPA 模板配置
FPA_TEMPLATE = {
    '需求名称': '关于集团事件业务影响分析开发需求',
    '需求编号': '49900_20250721',
    '版本': 'V3',
    '类型': '厂家版本',
    '统计日期': datetime.now().strftime('%Y-%m-%d')
}

# FPA Excel 列名映射 (标准格式)
FPA_COLUMNS = [
    '编号', '一级分类', '二级分类', '三级分类', '功能点名称', '功能点计数项',
    '类别', 'UFP', '重用程度', '修改类型', 'AFP', '备注'
]

# 工作量计算参数
WORK_PARAMS = {
    '规模变更调整因子': 1.21,  # CF
    '基准生产率': 10.12,  # PDR (人时/功能点)
    '调整因子': {
        '应用类型': 1.00,  # A
        '质量特性': 0.90,  # B
        '开发语言': 1.00,  # C
        '开发团队背景': 0.80  # D
    }
}


def clean_function_point_name(name: str) -> str:
    """
    清理功能点计数项名称，移除不适合的特殊符号
    
    规则：
    1. 将斜杠/替换为"和"字
    2. 移除中文引号（""）和英文引号（""）及内容
    3. 移除小括号（包括全角和半角）及内容
    4. 移除其他标点符号
    
    Args:
        name: 原始名称
    Returns:
        清理后的名称
    """
    import re
    
    # 特殊规则：将/替换为"和"字
    name = name.replace('/', '和')
    
    # 移除中文引号及内容："..."
    name = PATTERNS['note_zh'].sub('', name)
    
    # 移除英文引号及内容："..."
    name = PATTERNS['note_zh'].sub('', name)
    
    # 移除小括号及内容（全角和半角）
    name = PATTERNS['note_zh'].sub('', name)
    
    # 移除其他可能的特殊符号（保留中文、英文、数字、下划线、点号）
    # 只保留：中文、英文、数字、下划线、点号、空格
    name = PATTERNS['special_chars'].sub('', name)
    
    # 清理多余空格
    name = ' '.join(name.split())
    
    return name.strip()


def parse_requirement_document(md_content: str) -> list:
    """
    解析需求文档，提取 FPA 功能点信息
    
    Args:
        md_content: Markdown 格式的需求文档内容
        
    Returns:
        功能点列表
    """
    function_points = []
    
    # 清理零宽空格 (U+200B)
    md_content = PATTERNS['zero_width'].sub('', md_content)
    
    # 只解析"功能需求"章节之后的内容
    # 找到"# 功能需求"的位置
    func_req_index = md_content.find('# 功能需求')
    if func_req_index != -1:
        # 只截取"功能需求"章节之后的内容
        md_content = md_content[func_req_index:]
        logger.info(f"截取'功能需求'章节之后的内容进行解析")
    else:
        logger.warning("未找到'功能需求'章节，将解析全文")
    
    # 定义正则表达式模式（使用预编译的）
    # patterns = PATTERNS  # 已预编译
    
    lines = md_content.split('\n')
    current_point = {
        'level1': '',
        'level2': '',
        'level3': '',
        'level4': '',
        'level5': '',
        '功能描述': '',
        '系统界面': '',
        '输入': '',
        '输出': '',
        '处理过程': '',
        '内部逻辑文件数': 0,
        '外部逻辑文件数': 0,
        '新增/变更内部逻辑文件': '',
        '原有未修改内部逻辑文件': '',
        '新增/变更外部逻辑文件': '无',
        '原有未修改外部逻辑文件': '',
        # 新增字段用于标准 FPA 格式
        '类别': '',
        'UFP': 0,
        '重用程度': '',
        '修改类型': '',
        'AFP': 0,
        '备注': ''
    }
    
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # 跳过空行和注释
        if not line or line.startswith('<'):
            i += 1
            continue
        
        # 匹配各级标题（只匹配 level1-level5）
        for level_name in ['level1', 'level2', 'level3', 'level4', 'level5']:
            pattern = PATTERNS[level_name]
            match = pattern.match(line)
            if match:
                title = match.group(1).strip()
                # 去除标题中的加粗符号
                title = PATTERNS['bold'].sub('', title)
                
                if level_name == 'level5' and title:
                    # 保存上一个功能点
                    if current_point['level5']:
                        function_points.append(current_point.copy())
                                    
                    # 清洗标题中的注释和零宽字符
                    title_clean = PATTERNS['note_end'].sub('', title).strip()
                    title_clean = PATTERNS['zero_width'].sub('', title_clean).strip()
                    # 使用清理函数处理特殊符号
                    title_clean = clean_function_point_name(title_clean)
                                    
                    # 开始新的功能点 (同时清洗之前积累的 level1-4)
                    current_point = {
                        'level1': PATTERNS['category_note'].sub('', current_point['level1'].replace('\u200D', '')).strip(),
                        'level2': PATTERNS['category_note'].sub('', current_point['level2'].replace('\u200D', '')).strip(),
                        'level3': PATTERNS['category_note'].sub('', current_point['level3'].replace('\u200D', '')).strip(),
                        'level4': PATTERNS['category_note'].sub('', current_point['level4'].replace('\u200D', '')).strip(),
                        'level5': title_clean,
                        '功能点计数项': title_clean,  # 添加这个字段用于类别判断
                        '功能描述': '',
                        '系统界面': '',
                        '输入': '',
                        '输出': '',
                        '处理过程': '',
                        '内部逻辑文件数': 0,
                        '外部逻辑文件数': 0,
                        '新增/变更内部逻辑文件': '',
                        '原有未修改内部逻辑文件': '',
                        '新增/变更外部逻辑文件': '无',
                        '原有未修改外部逻辑文件': '',
                        # 新增字段用于标准 FPA 格式
                        '类别': '',
                        'UFP': 0,
                        '重用程度': '',
                        '修改类型': '',
                        'AFP': 0,
                        '备注': ''
                    }
                elif level_name == 'level4' and title:
                    current_point['level4'] = PATTERNS['category_note'].sub('', title.replace('\u200D', '')).strip()
                elif level_name == 'level3' and title:
                    current_point['level3'] = PATTERNS['category_note'].sub('', title.replace('\u200D', '')).strip()
                elif level_name == 'level2' and title:
                    current_point['level2'] = PATTERNS['category_note'].sub('', title.replace('\u200D', '')).strip()
                elif level_name == 'level1' and title:
                    current_point['level1'] = PATTERNS['category_note'].sub('', title.replace('\u200D', '')).strip()
                
                break
        
        # 提取功能描述字段
        if current_point['level5']:
            # 功能描述
            if line.startswith('**功能描述**') or line.startswith('**功能描述:**'):
                desc_match = PATTERNS['func_desc'].search(line)
                if desc_match:
                    current_point['功能描述'] = desc_match.group(1).strip()
            
            # 系统界面
            elif line.startswith('**系统界面**') or line.startswith('**系统界面:**'):
                desc_match = PATTERNS['func_desc'].search(line)
                if desc_match:
                    current_point['系统界面'] = desc_match.group(1).strip()
            
            # 输入
            elif line.startswith('**输入**') or line.startswith('**输入:**'):
                desc_match = PATTERNS['func_desc'].search(line)
                if desc_match:
                    current_point['输入'] = desc_match.group(1).strip()
            
            # 输出
            elif line.startswith('**输出**') or line.startswith('**输出:**'):
                desc_match = PATTERNS['func_desc'].search(line)
                if desc_match:
                    current_point['输出'] = desc_match.group(1).strip()
            
            # 处理过程
            elif line.startswith('**处理过程**') or line.startswith('**处理过程:**'):
                desc_match = PATTERNS['func_desc'].search(line)
                if desc_match:
                    current_point['处理过程'] = desc_match.group(1).strip()
            
            # 内部逻辑文件数量
            elif '本事务功能预计涉及到' in line and '个内部逻辑文件' in line:
                num_match = PATTERNS['ilf_count'].search(line)
                if num_match:
                    current_point['内部逻辑文件数'] = int(num_match.group(1))
                
                num_match = PATTERNS['elf_count'].search(line)
                if num_match:
                    current_point['外部逻辑文件数'] = int(num_match.group(1))
            
            # 新增/变更的内部逻辑文件
            elif '本期新增/变更的内部逻辑文件' in line:
                match = PATTERNS['ilf_files'].search(line)
                if match:
                    files = match.group(1).strip()
                    if files and files != '无':
                        current_point['新增/变更内部逻辑文件'] = files
            
            # 原有未修改的内部逻辑文件
            elif line.startswith('本期涉及原有但没修改的内部逻辑文件'):
                if ':' in line or ':' in line:
                    files = line.split(':', 1)[1].strip()
                    if files and files != '无':
                        current_point['原有未修改内部逻辑文件'] = files
            
            # 新增/变更的外部逻辑文件
            elif line.startswith('本期新增/变更的外部逻辑文件'):
                if ':' in line or ':' in line:
                    files = line.split(':', 1)[1].strip()
                    if files and files != '无':
                        current_point['新增/变更外部逻辑文件'] = files
            
            # 原有未修改的外部逻辑文件
            elif line.startswith('本期涉及原有但没修改的外部逻辑文件'):
                if ':' in line or ':' in line:
                    files = line.split(':', 1)[1].strip()
                    if files and files != '无':
                        current_point['原有未修改外部逻辑文件'] = files
        
        i += 1
    
    # 添加最后一个功能点
    if current_point['level5']:
        function_points.append(current_point)
        
    logger.info(f"解析完成，共提取 {len(function_points)} 个功能点")
            
    # 额外提取所有提到的内部逻辑文件 (表) 作为 ILF 功能点
    # 这是 FPA 标准格式的特殊要求
    # 关键规则：ILF 表应该按照它们在文档中出现的顺序插入到对应的功能点之后
    # ★★★★★ 重要：同一个 ILF 表名在整个需求中只能出现一次 ★★★★★
            
    # 全局去重：记录已经创建过的 ILF 表名
    created_ilf_tables = set()
    
    # 收集需要插入的 ILF 功能点，按插入位置分组
    ilf_insertions = {}  # {parent_index: [ilf_points]}
    
    for i, point in enumerate(function_points):
        ilf_files = point.get('新增/变更内部逻辑文件', '')
        if ilf_files and ilf_files != '无':
            # 分割多个表
            tables = [t.strip() for t in PATTERNS['table_split'].split(ilf_files)]
            valid_tables = [t for t in tables if t and t.endswith('表')]
            
            if not valid_tables:
                continue
            
            # 为每个表创建一个独立的 ILF 功能点
            for table_name in valid_tables:
                # 全局去重：如果这个表名已经创建过，跳过
                if table_name in created_ilf_tables:
                    logger.info(f"   跳过重复 ILF: {table_name} (已在前面创建)")
                    continue
                
                # 标记为已创建
                created_ilf_tables.add(table_name)
                
                # 创建 ILF 功能点
                new_point = point.copy()
                new_point['level5'] = table_name
                new_point['功能点计数项'] = table_name
                new_point['功能描述'] = f'{table_name}的维护与管理'
                new_point['新增/变更内部逻辑文件'] = table_name
                new_point['类别'] = 'ILF'  # 强制设为 ILF
                new_point['UFP'] = 7
                new_point['重用程度'] = '高'
                new_point['修改类型'] = '新增'
                new_point['AFP'] = round(7 * 0.33, 2)
                new_point['备注'] = f'提取自：{point.get("功能点计数项", "")}'
                
                # 记录插入位置（在当前功能点之后）
                if i not in ilf_insertions:
                    ilf_insertions[i] = []
                ilf_insertions[i].append(new_point)
                
                logger.info(f"  ✓ 创建 ILF: {table_name} (插入位置：功能点 {i} 之后)")
    
    # 从后往前插入 ILF 功能点，避免索引偏移
    if ilf_insertions:
        logger.info(f"\n开始插入 {sum(len(v) for v in ilf_insertions.values())} 个 ILF 功能点")
        for parent_idx in sorted(ilf_insertions.keys(), reverse=True):
            ilf_points = ilf_insertions[parent_idx]
            # 插入到父功能点之后
            insert_pos = parent_idx + 1
            for ilf_idx, ilf_point in enumerate(ilf_points):
                function_points.insert(insert_pos + ilf_idx, ilf_point)
            logger.info(f"  在功能点 {parent_idx} 之后插入 {len(ilf_points)} 个 ILF")
        
        logger.info(f"插入后总功能点数：{len(function_points)}\n")
    
    # 智能填充 FPA 字段 - 使用数据库配置的规则
    for point in function_points:
        item_text = point.get('功能点计数项', '')
        
        # 使用数据库中的规则判断类别
        try:
            category, ufp = FPACategoryRule.apply_rules(item_text)
            point['类别'] = category
            point['UFP'] = ufp
            logger.info(f"功能点 '{item_text}' -> 类别={category}, UFP={ufp}")
        except Exception as e:
            logger.error(f"应用类别规则失败：{e}，使用默认规则")
            # 降级到硬编码规则（如果数据库不可用）
            # 使用预编译的关键词模式加速匹配
            if KEYWORD_PATTERNS['ilf_suffix'].match(item_text) or KEYWORD_PATTERNS['ilf_keywords'].search(item_text):
                point['类别'] = 'ILF'
                point['UFP'] = 7
            elif KEYWORD_PATTERNS['ei_keywords'].search(item_text):
                point['类别'] = 'EI'
                point['UFP'] = 4
            elif KEYWORD_PATTERNS['eo_keywords'].search(item_text):
                point['类别'] = 'EO'
                point['UFP'] = 5
            elif KEYWORD_PATTERNS['eq_keywords'].search(item_text):
                point['类别'] = 'EQ'
                point['UFP'] = 4
            elif KEYWORD_PATTERNS['present_keywords'].search(item_text):
                if KEYWORD_PATTERNS['present_eq_keywords'].search(item_text):
                    point['类别'] = 'EO'
                    point['UFP'] = 5
                else:
                    point['类别'] = 'EQ'
                    point['UFP'] = 4
            else:
                point['类别'] = 'EO'
                point['UFP'] = 5
        
        # 2. 识别重用程度（全部设置为"高"）
        point['重用程度'] = '高'
        
        # 3. 修改类型 (全部设置为"新增")
        point['修改类型'] = '新增'
        
        # 4. 计算 AFP (根据重用程度和修改类型)
        ufp = point['UFP']
        reuse = point['重用程度']
        modify_type = point['修改类型']
        
        if modify_type == '新增':
            if reuse == '高':
                point['AFP'] = round(ufp * 0.33, 2)  # 高重用，33%
            elif reuse == '中':
                point['AFP'] = round(ufp * 0.67, 2)  # 中重用，67%
            else:
                point['AFP'] = ufp  # 低重用，100%
        else:
            point['AFP'] = round(ufp * 0.5, 2)  # 修改，50%
    
    return function_points


def sort_function_points(function_points: list) -> list:
    """
    按照指定的顺序对功能点进行排序
    
    Args:
        function_points: 功能点列表
        
    Returns:
        排序后的功能点列表
    """
    logger.info(f"=" * 80)
    logger.info(f"开始功能点排序验证")
    logger.info(f"=" * 80)
    
    # 读取顺序文件作为参考标准
    order_file_path = Path(__file__).parent.parent / 'test' / 'fpa' / '期望表格的顺序.txt'
    expected_order = []
    expected_categories = {}
    
    if order_file_path.exists():
        with open(order_file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            for line in lines[1:]:  # 跳过标题行
                parts = line.strip().split('\t')
                if len(parts) >= 2 and parts[0] != '必填':
                    item_name = parts[0].strip()
                    category = parts[1].strip() if len(parts) > 1 else ''
                    expected_order.append(item_name)
                    expected_categories[item_name] = category
        logger.info(f"从顺序文件读取了 {len(expected_order)} 个期望的功能点")
    else:
        logger.warning(f"顺序文件不存在：{order_file_path}")
        return function_points
    
    # 创建排序映射
    order_map = {name: idx for idx, name in enumerate(expected_order)}
    
    # 收集所有解析到的功能点名称
    parsed_names = [point.get('功能点计数项', '').strip() for point in function_points]
    
    logger.info(f"\n解析到的功能点总数：{len(parsed_names)}")
    logger.info(f"期望的功能点总数：{len(expected_order)}")
    
    # 详细比对
    logger.info(f"\n{'='*80}")
    logger.info(f"功能点匹配详情:")
    logger.info(f"{'='*80}")
    
    matched_count = 0
    unmatched_parsed = []
    unmatched_expected = []
    category_mismatches = []
    
    # 检查解析到的功能点
    for i, name in enumerate(parsed_names, 1):
        if name in order_map:
            matched_count += 1
            expected_idx = order_map[name]
            actual_category = next((p.get('类别', '') for p in function_points if p.get('功能点计数项', '').strip() == name), '')
            expected_category = expected_categories.get(name, '')
            
            if actual_category != expected_category and expected_category:
                category_mismatches.append({
                    'name': name,
                    'actual': actual_category,
                    'expected': expected_category
                })
                logger.info(f"[{i}/{len(parsed_names)}] ✓ {name} - 类别不匹配！实际：{actual_category}, 期望：{expected_category}")
            else:
                logger.info(f"[{i}/{len(parsed_names)}] ✓ {name} - 匹配 (位置：{expected_idx}, 类别：{actual_category})")
        else:
            unmatched_parsed.append(name)
            logger.info(f"[{i}/{len(parsed_names)}] ✗ {name} - 未在期望列表中找到!")
    
    # 检查期望的功能点是否有遗漏
    for name in expected_order:
        if name not in parsed_names:
            unmatched_expected.append(name)
            logger.info(f"✗ 期望但缺失：{name}")
    
    # 汇总统计
    logger.info(f"\n{'='*80}")
    logger.info(f"匹配统计:")
    logger.info(f"{'='*80}")
    logger.info(f"成功匹配：{matched_count}/{len(parsed_names)} ({matched_count/len(parsed_names)*100:.1f}%)")
    logger.info(f"解析但未期望：{len(unmatched_parsed)} 个")
    logger.info(f"期望但未解析：{len(unmatched_expected)} 个")
    logger.info(f"类别不匹配：{len(category_mismatches)} 个")
    
    if unmatched_parsed:
        logger.info(f"\n解析但未在期望列表中的功能点:")
        for name in unmatched_parsed:
            logger.info(f"  - {name}")
    
    if unmatched_expected:
        logger.info(f"\n期望但未解析到的功能点:")
        for name in unmatched_expected:
            logger.info(f"  - {name}")
    
    if category_mismatches:
        logger.info(f"\n类别不匹配的功能点:")
        for mismatch in category_mismatches:
            logger.info(f"  - {mismatch['name']}: 实际={mismatch['actual']}, 期望={mismatch['expected']}")
    
    logger.info(f"\n{'='*80}")
    
    # 排序功能点
    def get_sort_key(point):
        item_name = point.get('功能点计数项', '').strip()
        # 如果在预期列表中，返回对应的索引
        if item_name in order_map:
            return order_map[item_name]
        # 如果不在预期列表中，返回一个很大的数，排在最后
        return len(expected_order)
    
    sorted_points = sorted(function_points, key=get_sort_key)
    
    logger.info(f"功能点排序完成：共 {len(sorted_points)} 个功能点")
    logger.info(f"排序后的前 10 个功能点:")
    for i, point in enumerate(sorted_points[:10], 1):
        logger.info(f"  {i}. {point.get('功能点计数项', '')} (类别：{point.get('类别', '')})")
    
    return sorted_points


def generate_fpa_excel(function_points: list, output_path: str) -> str:
    """
    生成 FPA预估 Excel 文件 (标准格式)
    
    Args:
        function_points: 功能点列表
        output_path: 输出文件路径
        
    Returns:
        生成的文件路径
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # 创建工作簿
    wb = Workbook()
    
    # 删除默认 sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 1. 先创建"1. 填写说明"sheet
    ws_info = wb.create_sheet(title='1. 填写说明')
    
    # 样式定义 (移到这里，确保在使用前定义)
    header_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")  # 灰色表头
    header_font = Font(bold=True, size=11, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色 (被保护)
    light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 浅绿色
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置"1. 填写说明"的内容
    ws_info.merge_cells('A1:B1')
    ws_info['A1'] = '模型填写顺序为：\n1、拆分功能点，填写《2.规模估算》。\n2、调整因子参考《3.调整因子》进行取值。\n3、自动计算评估结果，查看《4.评估结果》。'
    ws_info['A1'].alignment = Alignment(horizontal='left', vertical='top')
    ws_info.row_dimensions[1].height = 100
    
    ws_info['A2'] = '白色单元格'
    ws_info['B2'] = '计价评估时，只填写白色单元格'
    ws_info['A3'] = '深灰单元格'
    ws_info['B3'] = '模板格式部分，不得修改'
    ws_info['A4'] = '绿色单元格'
    ws_info['B4'] = '公式计算结果，不得擅自修改'
    ws_info['A5'] = '红色字体'
    ws_info['B5'] = '公式计算结果或说明'
    
    for row in range(1, 6):
        for col in range(1, 3):
            cell = ws_info.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 50
    
    # 2. 创建"2. 规模估算"sheet
    ws = wb.create_sheet(title='2. 规模估算')
    
    # 3. 创建"3. 调整因子"sheet
    ws3 = wb.create_sheet(title='3. 调整因子')
    
    # 设置"3. 调整因子"的标题
    ws3.merge_cells('A1:C1')
    ws3['A1'] = '3. 调整因子列表'
    ws3['A1'].font = Font(bold=True, size=16)
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 40
    
    # 表头 (第 2 行)
    ws3['A2'] = '因子类型'
    ws3['B2'] = '因子名称'
    ws3['C2'] = '因子计算结果'
    
    # 设置表头样式
    for col in ['A', 'B', 'C']:
        cell = ws3.cell(row=2, column=ord(col) - ord('A') + 1)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    ws3.row_dimensions[2].height = 30
    
    # 完整的调整因子数据（按照参考 Excel 格式）
    factor_data = [
        # 规模计数时机
        ('规模计数时机', '估算中期', "=IF('3. 调整因子'!B3='3. 调整因子'!B37,'3. 调整因子'!D37,IF('3. 调整因子'!B3='3. 调整因子'!B38,'3. 调整因子'!D38,IF('3. 调整因子'!B3='3. 调整因子'!B39,'3. 调整因子'!D39,IF('3. 调整因子'!B3='3. 调整因子'!B40,'3. 调整因子'!D40,1.21))))"),
        # 应用类型
        ('应用类型', '业务处理', "=IF(B4=B13,C13,IF(B4=B14,C14,IF(B4=B15,C15,IF(B4=B16,C16,IF(B4=B17,C17,IF(B4=B18,C18,IF(B4=B19,C19,IF(B4=B20,C20,IF(B4=B21,C21,1)))))))))"),
        # 质量特性 - 分布式处理
        ('质量特性 - 分布式处理', '没有明示对分布式处理的需求事项', "=IF(B5=C23,D23,IF(B5=C24,D24,IF(B5=C25,D25,-10)))"),
        # 质量特性 - 性能
        ('质量特性 - 性能', '没有明示对性能的特别需求事项或活动，因此提供基本性能', "=IF(B6=C26,D26,IF(B6=C27,D27,IF(B6=C28,D28,-10)))"),
        # 质量特性 - 可靠性
        ('质量特性 - 可靠性', '没有明示对可靠性的特别需求事项或活动，因此提供基本的可靠性', "=IF(B7=C29,D29,IF(B7=C30,D30,IF(B7=C31,D31,-10)))"),
        # 质量特性 - 多重站点
        ('质量特性 - 多重站点', '在相同用途的硬件或软件环境下运行', "=IF(B8=C32,D32,IF(B8=C33,D33,IF(B8=C34,D34,-10)))"),
        # 开发语言
        ('开发语言', 'JAVA、C++、C#及其他同级别语言/平台', "=IF(B9=B44,C44,IF(B9=B45,C45,IF(B9=B46,C46,1)))"),
        # 开发团队背景
        ('开发团队背景', '为本行业（政府）开发过类似的软件', "=IF(B10=B49,C49,IF(B10=B50,C50,IF(B10=B51,C51,0.8)))"),
        # 空行
        (None, None, None),
        # 应用类型详细说明
        ('应用类型', '描述', '调整因子'),
        (None, '业务处理', '办公自动化系统：人事、会计、工资、销售等经营管理及业务处理用软件等', 1),
        (None, '应用集成', '企业服务总线、应用集成等', 1.2),
        (None, '科技', '科学计算、仿真、基于复杂算法的统计分析等', 1.2),
        (None, '多媒体', '图形、影像、声音等多媒体应用领域：地理信息系统：教育和娱乐等', 1.5),
        (None, '智能信息', '自然语言处理、大模型、计算机视觉、智能决策、专家系统等', 1.7),
        (None, '基础软件/支撑软件', '操作系统、数据库系统、集成开发环境、自动化开发/设计工具等', 1.7),
        (None, '通信控制', '通信协议、仿真、交换机软件、全球定位系统等', 1.9),
        (None, '流程控制', '实时系统控制、机器人控制、嵌入式软件等', 2),
        # 空行
        (None, None, None),
        # 质量特性详细说明
        ('调整因子', '判断标准', '调整因子'),
        (None, '分布式处理', '没有明示对分布式处理的需求事项', -1),
        (None, None, '通过网络进行客户端/服务器及网络基础应用分布处理和传输', 0),
        (None, None, '在多个服务器及处理器上同时相互执行计算机系统中的处理功能', 1),
        (None, '性能', '没有明示对性能的特别需求事项或活动，因此提供基本性能', -1),
        (None, None, '应答时间或处理率对高峰时间或所有业务时间来说都很重要，对连动系统结束处理时间的限制', 0),
        (None, None, '为满足性能需求事项，要求设计阶段开始进行性能分析，或在设计、开发阶段使用分析工具', 1),
        (None, '可靠性', '没有明示对可靠性的特别需求事项或活动，因此提供基本的可靠性', -1),
        (None, None, '发生故障时可轻易修复，带来一定不便或经济损失', 0),
        (None, None, '发生故障时很难复，发生重大经济损失或有生命危害', 1),
        (None, '多重站点', '在相同用途的硬件或软件环境下运行', -1),
        (None, None, '在用途类似的硬件或软件环境下运行', 0),
        (None, None, '在不同用途的硬件或软件环境下运行', 1),
        # 空行
        (None, None, None),
        # 规模变更调整因子 (CF)
        (None, '规模变更调整因子 (CF)', None),
        (None, '估算早期', '概算、预算阶段', 1.39),
        (None, '估算中期', '投标、项目计划阶段', 1.21),
        (None, '估算晚期', '需求分析阶段', 1.1),
        (None, '项目完成', '项目交付后及运维阶段', 1),
        # 空行
        (None, None, None),
        # 开发语言详细说明
        (None, '开发语言', None, '调整因子'),
        (None, 'C 及其他同级别语言/平台', None, 1.2),
        (None, 'JAVA、C++、C#及其他同级别语言/平台', None, 1),
        (None, 'PowerBuilder、ASP 及其他同级别语言/平台', None, 0.8),
        # 空行
        (None, None, None),
        # 开发团队背景详细说明
        (None, '开发团队背景', None, '调整因子'),
        (None, '为本行业（政府）开发过类似的软件', None, 0.8),
        (None, '为其他行业开发过类似的软件，或为本行业（政府）开发过不同但相关的软件', None, 1),
        (None, '未开发过类似软件', None, 1.2),
    ]
    
    row_idx = 3
    for data in factor_data:
        if len(data) == 3:
            col_a, col_b, col_c = data
            col_d = None
        else:
            col_a, col_b, col_c, col_d = data
            
        if col_a is not None:
            ws3.cell(row=row_idx, column=1, value=col_a).alignment = Alignment(horizontal='left', vertical='center')
        if col_b is not None:
            ws3.cell(row=row_idx, column=2, value=col_b).alignment = Alignment(horizontal='left', vertical='center')
        if col_c is not None:
            ws3.cell(row=row_idx, column=3, value=col_c).alignment = Alignment(horizontal='left', vertical='center')
        if col_d is not None:
            ws3.cell(row=row_idx, column=4, value=col_d).alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置边框
        for col in range(1, 5):
            cell = ws3.cell(row=row_idx, column=col)
            cell.border = thin_border
        
        row_idx += 1
    
    # 设置列宽
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 50
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 12
    # E~I 列设置为空但需要存在
    for col in ['E', 'F', 'G', 'H', 'I']:
        ws3.column_dimensions[col].width = 10

    # 1. 标题
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = '2. 规模估算'
    title_cell.font = Font(bold=True, size=20)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # 2. 规模估算方法 (第 3-5 行)
    ws.merge_cells('A3:B3')
    ws['A3'] = '规模估算方法'
    ws['A3'].font = Font(bold=True)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A3'].fill = header_fill

    ws.merge_cells('C3:L3')
    ws['C3'] = "'软件开发计价模型':7/5/4/5/4"
    ws['C3'].alignment = Alignment(horizontal='left', vertical='center')
    ws['C3'].font = Font(color="FF0000")  # 红色字体

    # 3. 未调整功能点合计
    ws['A4'] = '未调整功能点合计'
    ws['A4'].font = Font(bold=True)
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')

    # 计算 UFP 总和
    total_ufp = sum(point.get('UFP', 7) for point in function_points)
    ws['B4'] = total_ufp
    ws['B4'].fill = green_fill
    ws['B4'].font = Font(bold=True, color="0000FF")  # 蓝色字体

    ws.merge_cells('C4:L4')
    ws['C4'] = 'UFP,单位:FP'
    ws['C4'].font = Font(color="FF0000")

    # 4. 调整后功能点合计
    ws['A5'] = '调整后功能点合计'
    ws['A5'].font = Font(bold=True)
    ws['A5'].alignment = Alignment(horizontal='center', vertical='center')

    # 计算 AFP 总和
    total_afp = sum(point.get('AFP', 0) for point in function_points)
    ws['B5'] = round(total_afp, 2)
    ws['B5'].fill = green_fill
    ws['B5'].font = Font(bold=True, color="0000FF")

    ws.merge_cells('C5:L5')
    ws['C5'] = 'AFP,单位:FP'
    ws['C5'].font = Font(color="FF0000")

    # 5. 说明
    ws.merge_cells('A6:L6')
    ws['A6'] = '1. 本次计数在项目前期，需求未充分挖掘'
    ws['A6'].font = Font(italic=True)
    ws.row_dimensions[6].height = 25

    ws.merge_cells('A7:L7')
    ws['A7'] = '2. 绿色单元格为被保护，不得擅自修改'
    ws['A7'].font = Font(italic=True)
    ws['A7'].font = Font(color="008000")  # 绿色字体
    ws.row_dimensions[7].height = 25

    # 6. 表头 (第 8-9 行)
    headers = ['编号', '一级分类', '二级分类', '三级分类', '功能点名称', '功能点计数项',
               '类别', 'UFP', '重用程度', '修改类型', 'AFP', '备注']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.row_dimensions[8].height = 40

    # 7. 填写说明 (第 9 行)
    instructions = ['填写说明', '必填', '必填', '必填', '必填', '必填',
                   '请选择计价模型种类。', '自动计算未调整的功能点数量。', '请选择代码复用情况', '请选择代码修改情况',
                   '自动计算调整后的功能点数量。', '调整 AFP:请通过重用程度、修改类型来改']

    for col, instruction in enumerate(instructions, 1):
        cell = ws.cell(row=9, column=col, value=instruction)
        cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  # 橙色
        cell.font = Font(size=9, color="FF0000")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.row_dimensions[9].height = 60

    # 8. 数据行
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for row_idx, point in enumerate(function_points, 10):
        # 编号
        ws.cell(row=row_idx, column=1, value=row_idx - 9).alignment = Alignment(horizontal='center')

        # 分类信息 (去除所有注释文字)
        level1 = re.sub(r'\s*[（(]注.*?[)）]\s*$', '', point.get('level1', '')).strip()
        level2 = re.sub(r'\s*[（(]注.*?[)）]\s*$', '', point.get('level2', '')).strip()
        level3 = re.sub(r'\s*[（(]注.*?[)）]\s*$', '', point.get('level3', '')).strip()
        level4 = re.sub(r'\s*[（(]注.*?[)）]\s*$', '', point.get('level4', '')).strip()
        level5 = re.sub(r'\s*[（(]注.*?[)）]\s*$', '', point.get('level5', '')).strip()

        ws.cell(row=row_idx, column=2, value=level1).alignment = data_alignment
        ws.cell(row=row_idx, column=3, value=level2).alignment = data_alignment
        ws.cell(row=row_idx, column=4, value=level3).alignment = data_alignment
        ws.cell(row=row_idx, column=5, value=level4).alignment = data_alignment
        ws.cell(row=row_idx, column=6, value=level5).alignment = data_alignment

        # 类别 (EI/EO/EQ/ILF/EIF)
        category = point.get('类别', 'EO')
        ws.cell(row=row_idx, column=7, value=category).alignment = Alignment(horizontal='center')

        # UFP (使用公式计算 - 根据类别自动计算)
        # ILF=7, EO=5, EIF=5, EI=4, EQ=4
        ufp_cell = ws.cell(row=row_idx, column=8)
        ufp_cell.value = f'=IF(G{row_idx}="ILF",7,IF(G{row_idx}="EO",5,IF(G{row_idx}="EIF",5,IF(G{row_idx}="EI",4,IF(G{row_idx}="EQ",4,5)))))'
        ufp_cell.fill = green_fill
        ufp_cell.alignment = Alignment(horizontal='center')

        # 重用程度（确保有默认值）
        reuse = point.get('重用程度') or '高'
        ws.cell(row=row_idx, column=9, value=reuse).alignment = Alignment(horizontal='center')

        # 修改类型（确保有默认值）
        modify_type = point.get('修改类型') or '新增'
        ws.cell(row=row_idx, column=10, value=modify_type).alignment = Alignment(horizontal='center')

        # AFP (使用公式计算 - 根据重用程度和修改类型)
        # 新增：高重用 33%, 中重用 67%, 低重用 100%
        # 修改：50%
        afp_cell = ws.cell(row=row_idx, column=11)
        afp_cell.value = f'=IF(J{row_idx}="新增",IF(I{row_idx}="高",H{row_idx}*0.33,IF(I{row_idx}="中",H{row_idx}*0.67,H{row_idx})),H{row_idx}*0.5)'
        afp_cell.fill = green_fill
        afp_cell.alignment = Alignment(horizontal='center')

        # 备注
        ws.cell(row=row_idx, column=12, value=point.get('备注', '')).alignment = data_alignment

        # 设置所有单元格边框
        for col in range(1, 13):
            ws.cell(row=row_idx, column=col).border = thin_border

        ws.row_dimensions[row_idx].height = 30

    # 9. 合计行
    last_row = len(function_points) + 10
    ws.merge_cells(f'A{last_row}:G{last_row}')
    summary_cell = ws.cell(row=last_row, column=1, value='合计')
    summary_cell.font = Font(bold=True, size=12)
    summary_cell.alignment = Alignment(horizontal='right')

    ws.cell(row=last_row, column=8, value=total_ufp).font = Font(bold=True)
    ws.cell(row=last_row, column=11, value=round(total_afp, 2)).font = Font(bold=True)

    for col in range(1, 13):
        ws.cell(row=last_row, column=col).border = thin_border

    # 设置列宽
    column_widths = {
        'A': 6,   # 编号
        'B': 15,  # 一级分类
        'C': 15,  # 二级分类
        'D': 15,  # 三级分类
        'E': 20,  # 功能点名称
        'F': 40,  # 功能点计数项
        'G': 10,  # 类别
        'H': 8,   # UFP
        'I': 10,  # 重用程度
        'J': 10,  # 修改类型
        'K': 8,   # AFP
        'L': 30   # 备注
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 10. 创建"4. 评估结果"sheet
    ws2 = wb.create_sheet(title='4. 评估结果')
    
    # 标题
    ws2.merge_cells('A1:C1')
    title_cell = ws2['A1']
    title_cell.value = '4. 评估结果'
    title_cell.font = Font(bold=True, size=20)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 40
    
    # 评估结果数据
    cf = WORK_PARAMS['规模变更调整因子']
    pdr = WORK_PARAMS['基准生产率']
    
    factors = WORK_PARAMS['调整因子']
    
    # 表头 (第 3 行)
    ws2['C3'] = '数值'
    ws2['D3'] = '说明'
    ws2['C3'].font = Font(bold=True)
    ws2['D3'].font = Font(bold=True)
    ws2['C3'].fill = header_fill
    ws2['D3'].fill = header_fill
    
    # 第 4 行 - 规模估算结果 (A4+B4 合并)
    ws2.merge_cells('A4:B4')
    ws2['A4'] = '规模估算结果 (单位：功能点)'
    ws2['C4'] = f"=SUM('2. 规模估算'!K10:K{last_row-1})"  # 动态计算 2.规模估算表中 K 列 (AFP) 的总和
    ws2['D4'] = 'AFP'
    
    # 第 5 行 - 规模变更调整因子 (A5+B5 合并)
    ws2.merge_cells('A5:B5')
    ws2['A5'] = '规模变更调整因子'
    ws2['C5'] = "='2. 规模估算'!C5"  # 引用 2.规模估算表的 C5
    ws2['D5'] = 'CF, 项目阶段'
    
    # 第 6 行 - 调整后规模 (A6+B6 合并)
    ws2.merge_cells('A6:B6')
    ws2['A6'] = '调整后规模 (单位：功能点)'
    ws2['C6'] = '=C4*C5'  # 公式计算
    ws2['D6'] = 'S, 等于 AFP*CF'
    
    # 第 7 行 - 基准生产率 (A7+B7 合并)
    ws2.merge_cells('A7:B7')
    ws2['A7'] = '基准生产率 (单位：人时/功能点)'
    ws2['C7'] = pdr
    ws2['D7'] = 'PDR, 来自《2024 年软件行业基准数据》'
    
    # 第 8 行 - 未调整工作量 (A8+B8 合并)
    ws2.merge_cells('A8:B8')
    ws2['A8'] = '未调整工作量 (单位：人天)'
    ws2['C8'] = '=C6*C7/8'  # 公式计算
    ws2['D8'] = 'S*PDR/8h'
    
    # 第 9-12 行 - 调整因子 (A9:A12 合并)
    ws2.merge_cells('A9:A12')
    ws2['A9'] = '调整因子'
    ws2['A9'].font = Font(bold=True)
    ws2['A9'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 第 9 行 - 应用类型
    ws2['B9'] = '应用类型'
    ws2['C9'] = factors['应用类型']
    ws2['D9'] = "='3. 调整因子'!D3"  # A,应用类型调整因子
    
    # 第 10 行 - 质量特性
    ws2['B10'] = '质量特性'
    ws2['C10'] = factors['质量特性']
    ws2['D10'] = '=1+0.025*SUM(\'3. 调整因子\'!D4:D7)'  # B,质量特性调整因子
    
    # 第 11 行 - 开发语言
    ws2['B11'] = '开发语言'
    ws2['C11'] = factors['开发语言']
    ws2['D11'] = "='3. 调整因子'!D8"  # C,开发语言调整因子
    
    # 第 12 行 - 开发团队背景
    ws2['B12'] = '开发团队背景'
    ws2['C12'] = factors['开发团队背景']
    ws2['D12'] = "='3. 调整因子'!D9"  # D,开发团队背景调整因子
    
    # 第 13 行 - 调整后工作量 (A13+B13 合并)
    ws2.merge_cells('A13:B13')
    ws2['A13'] = '调整后工作量 (单位：人天)'
    ws2['C13'] = '=C8*C9*C10*C11*C12'  # 公式计算
    ws2['D13'] = 'AE,S*PDR/8h*A*B*C*D'
    
    # 设置样式
    for row in range(4, 14):
        for col in range(1, 5):
            cell = ws2.cell(row=row, column=col)
            cell.border = thin_border
    
    # 绿色单元格 (被保护)
    green_cells = ['C4', 'C5', 'C6', 'C8', 'C9', 'C10', 'C11', 'C12', 'C13']
    for cell_ref in green_cells:
        col = cell_ref[0].upper()
        row = int(cell_ref[1:])
        cell = ws2.cell(row=row, column=ord(col)-ord('A')+1)
        cell.fill = green_fill
        cell.font = Font(bold=True, color="0000FF")
    
    # 红色说明
    red_cells = ['D7', 'D13']
    for cell_ref in red_cells:
        col = cell_ref[0].upper()
        row = int(cell_ref[1:])
        cell = ws2.cell(row=row, column=ord(col)-ord('A')+1)
        cell.font = Font(color="FF0000")

    # 设置评估结果表列宽
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 40
    
    # 保存文件
    wb.save(output_path)
    logger.info(f"FPA Excel 生成成功：{output_path}")
    return output_path


# ---------------------- 路由定义 ----------------------

@fpa_generator_bp.route('/')
def fpa_generator_page():
    """FPA 预估表生成主页面"""
    return render_template('fpa/fpa_generator.html')


@fpa_generator_bp.route('/upload', methods=['POST'])
def upload_requirement():
    """
    上传需求文档并生成 FPA 预估表
    支持 Markdown (.md) 和 Word (.docx) 格式
    
    Form Data:
        requirement_file: Markdown 或 Word 格式的需求文档
    """
    try:
        if 'requirement_file' not in request.files:
            flash("未选择上传文件！", "error")
            return render_template('fpa/fpa_generator.html')
        
        file = request.files['requirement_file']
        if file.filename == '':
            flash("未选择有效的文件！", "error")
            return render_template('fpa/fpa_generator.html')
        
        # 检查文件类型
        filename_lower = file.filename.lower()
        is_word_file = filename_lower.endswith('.docx')
        is_md_file = filename_lower.endswith('.md')
        
        logger.info(f"上传文件：{file.filename}, 是否 Word: {is_word_file}, 是否 MD: {is_md_file}")
        
        if not (is_word_file or is_md_file):
            logger.error(f"不支持的文件类型：{file.filename}")
            flash(f"请上传 Markdown (.md) 或 Word (.docx) 文件！当前文件：{file.filename}", "error")
            return render_template('fpa/fpa_generator.html')
        
        # 保存临时文件
        timestamp = int(time.time() * 1000)
        upload_dir = Path(current_app.config['UPLOAD_FOLDER']) / "fpa_input"
        output_dir = Path(current_app.config['UPLOAD_FOLDER']) / "fpa_output"
        os.makedirs(upload_dir, exist_ok=True)
        os.makedirs(output_dir, exist_ok=True)
        
        filename = Path(file.filename).stem
        temp_file_path = upload_dir / f"{filename}_{timestamp}{'.' + file.filename.rsplit('.', 1)[1].lower()}"
        file.save(temp_file_path)
        
        # 如果是 Word 文件，先转换为 Markdown
        if is_word_file:
            logger.info(f"检测到 Word 文件，开始转换为 Markdown: {temp_file_path}")
            try:
                from docx import Document
                from markdownify import markdownify as md
                
                # 读取 Word 文档
                doc = Document(temp_file_path)
                md_lines = []
                
                for para in doc.paragraphs:
                    style = para.style.name.lower()
                    text = ''.join(
                        f'**{run.text}**' if run.bold else 
                        f'*{run.text}*' if run.italic else 
                        run.text
                        for run in para.runs
                    ).strip()
                    
                    if not text:
                        continue
                    
                    # 根据样式转换标题
                    if 'heading 1' in style:
                        md_lines.append(f"# {text}")
                    elif 'heading 2' in style:
                        md_lines.append(f"## {text}")
                    elif 'heading 3' in style:
                        md_lines.append(f"### {text}")
                    elif 'heading 4' in style:
                        md_lines.append(f"#### {text}")
                    elif 'heading 5' in style:
                        md_lines.append(f"##### {text}")
                    elif 'heading 6' in style:
                        md_lines.append(f"###### {text}")
                    else:
                        md_lines.append(text)
                
                md_content = '\n\n'.join(md_lines)
                
                # 保存为 Markdown 文件
                temp_md_path = upload_dir / f"{filename}_{timestamp}.md"
                with open(temp_md_path, 'w', encoding='utf-8') as f:
                    f.write(md_content)
                
                # 删除临时 Word 文件
                os.remove(temp_file_path)
                temp_file_path = temp_md_path
                
                logger.info(f"Word 转 Markdown 成功：{temp_md_path}")
                
            except Exception as e:
                logger.error(f"Word 转 Markdown 失败：{e}", exc_info=True)
                flash(f"Word 文件转换失败：{str(e)}", "error")
                return render_template('fpa/fpa_generator.html')
        else:
            temp_md_path = temp_file_path
        
        # 读取并解析 Markdown 文档
        with open(temp_md_path, 'r', encoding='utf-8') as f:
            md_content = f.read()
        
        function_points = parse_requirement_document(md_content)
        
        if not function_points:
            flash("未能从文档中提取到功能点信息！", "error")
            return render_template('fpa/fpa_generator.html')
        
        # 从数据库读取评估结果中的 AFP 目标值
        try:
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT afp, adjusted_scale FROM fpa_evaluation_result
                ORDER BY created_at DESC
                LIMIT 1
            """)
            eval_result = cursor.fetchone()
            cursor.close()
            conn.close()
                            
            if eval_result and eval_result.get('afp'):
                # 获取目标 AFP 值（来自 Excel 的"规模估算结果 (单位：功能点)"）
                target_afp_from_excel = float(eval_result['afp'])
                cf = WORK_PARAMS.get('规模变更调整因子', 1.21)
                            
                # 计算当前解析的功能点 AFP 总和
                current_afp = sum(point.get('AFP', 0) for point in function_points)
                            
                # 计算 AFP 差值和扩展需求
                afp_difference = target_afp_from_excel - current_afp
                need_ai_expansion = current_afp < target_afp_from_excel
                            
                # 详细日志输出
                logger.info("=" * 80)
                logger.info("📊 AFP 数据对比分析")
                logger.info("=" * 80)
                logger.info(f"【1】从数据库读取的目标 AFP 值:")
                logger.info(f"    - 来源：fpa_evaluation_result 表最新记录")
                logger.info(f"    - target_afp_from_excel = {target_afp_from_excel:.2f}")
                logger.info(f"")
                logger.info(f"【2】当前生成的功能点 AFP 总和:")
                logger.info(f"    - 功能点总数：{len(function_points)}个")
                logger.info(f"    - current_afp = Σ(每个功能点的 AFP) = {current_afp:.2f}")
                logger.info(f"")
                logger.info(f"【3】AFP 差值计算:")
                logger.info(f"    - afp_difference = target_afp_from_excel - current_afp")
                logger.info(f"    - afp_difference = {target_afp_from_excel:.2f} - {current_afp:.2f}")
                logger.info(f"    - afp_difference = {afp_difference:.2f}")
                logger.info(f"")
                logger.info(f"【4】AI 扩展判断:")
                logger.info(f"    - 判断条件：current_afp < target_afp_from_excel")
                logger.info(f"    - 判断公式：{current_afp:.2f} < {target_afp_from_excel:.2f}")
                logger.info(f"    - need_ai_expansion = {need_ai_expansion}")
                logger.info(f"")
                            
                if need_ai_expansion:
                    logger.info(f"✅ 结论：需要 AI 扩展（当前 AFP {current_afp:.2f} < 目标 AFP {target_afp_from_excel:.2f}）")
                    logger.info(f"   AFP 缺口：{afp_difference:.2f}")
                                
                    # 计算需要扩展的功能点数量
                    avg_afp_per_point = 1.65  # UFP=5 × 0.33（高重用新增功能点）
                    needed_expand_count = int((afp_difference / avg_afp_per_point) + 0.5)
                                
                    logger.info(f"")
                    logger.info(f"【5】扩展功能点数量估算:")
                    logger.info(f"    - 假设每功能点平均 AFP: {avg_afp_per_point:.2f} (UFP=5 × 0.33)")
                    logger.info(f"    - 计算公式：needed_expand_count = afp_difference / avg_afp_per_point")
                    logger.info(f"    - needed_expand_count = {afp_difference:.2f} / {avg_afp_per_point:.2f}")
                    logger.info(f"    - needed_expand_count = {needed_expand_count}个")
                    logger.info(f"")
                    logger.info(f"   🎯 目标：通过 AI 扩展约 {needed_expand_count} 个功能点，使总 AFP 达到或超过 {target_afp_from_excel:.2f}")
                    logger.info("=" * 80)
                                
                    # 尝试使用 AI 辅助生成新的功能点
                    try:
                        logger.info("")
                        logger.info("🤖 [AI_EXPAND] 开始调用 AI 模型进行功能点拆分...")
                        logger.info(f"[AI_EXPAND] 扩展目标：{needed_expand_count}个功能点")
                        logger.info(f"[AI_EXPAND] AFP 目标：弥补 {afp_difference:.2f} 的差距")
                        logger.info(f"[AI_EXPAND] 参考功能点：{len(function_points)}个")
                        
                        # 配置项：是否使用 OMLX 模型（Qwen3.5-4B-OptiQ-4bit）
                        # True = 使用 OMLX 在线模型，False = 使用本地 Ollama 模型
                        USE_OMLX = True  # 修改这里来切换模型：True=OMLX, False=本地 Ollama
                        
                        if USE_OMLX:
                            logger.info("[AIC_EXPAND] 🌐 使用 OMLX 模型：Qwen3.5-4B-OptiQ-4bit")
                        else:
                            logger.info("[AI_EXPAND] 💻 使用本地 Ollama 模型：qwen3:4b")
                                
                        # 从原始功能点中选择有代表性的进行拆分
                        expanded_points = ai_assisted_expand_function_points(
                            function_points,  # 传入所有功能点作为参考
                            needed_expand_count,
                            use_omlx=USE_OMLX  # 传递模型选择参数
                        )
                                                    
                        if expanded_points:
                            # 关键修复：将扩展的功能点插入到对应原始功能点之后，保持文档顺序
                            # 1. 首先给每个原始功能点添加索引标记
                            for idx, point in enumerate(function_points):
                                point['_original_index'] = idx
                            
                            # 2. 按原始索引分组扩展的功能点
                            from collections import defaultdict
                            expanded_by_parent = defaultdict(list)
                            for exp_point in expanded_points:
                                parent_idx = exp_point.get('_parent_index', -1)
                                expanded_by_parent[parent_idx].append(exp_point)
                            
                            # 3. 从后向前插入，避免索引偏移问题
                            # （从大到小排序，这样插入时不会影响前面的索引）
                            sorted_indices = sorted(expanded_by_parent.keys(), reverse=True)
                            
                            for parent_idx in sorted_indices:
                                children = expanded_by_parent[parent_idx]
                                if parent_idx < len(function_points):
                                    # 在原始功能点之后插入所有子功能点
                                    insert_pos = parent_idx + 1
                                    for child in children:
                                        function_points.insert(insert_pos, child)
                                        insert_pos += 1
                                    logger.debug(f"在索引 {parent_idx} 后插入 {len(children)} 个子功能点")
                            
                            # 4. 清理临时索引字段
                            for point in function_points:
                                if '_original_index' in point:
                                    del point['_original_index']
                                if '_parent_index' in point:
                                    del point['_parent_index']
                                                    
                            # 重新计算扩展后的 AFP 总和
                            new_afp_total = sum(point.get('AFP', 0) for point in function_points)
                            afp_increase = new_afp_total - current_afp
                            actual_avg_afp = afp_increase / len(expanded_points) if expanded_points else 0
                                                    
                            logger.info(f"")
                            logger.info("✅ [AI_EXPAND] AI 扩展成功！详细统计:")
                            logger.info(f"  【1】扩展功能点数量:")
                            logger.info(f"      - 扩展前功能点总数：{len(function_points) - len(expanded_points)}个")
                            logger.info(f"      - 扩展后功能点总数：{len(function_points)}个")
                            logger.info(f"      - 新增功能点数量：{len(expanded_points)}个")
                            logger.info(f"")
                            logger.info(f"  【2】AFP 变化统计:")
                            logger.info(f"      - 扩展前 AFP 总和：{current_afp:.2f}")
                            logger.info(f"      - 扩展后 AFP 总和：{new_afp_total:.2f}")
                            logger.info(f"      - AFP 增加量：{afp_increase:.2f}")
                            logger.info(f"      - 实际平均每功能点 AFP: {actual_avg_afp:.2f}")
                            logger.info(f"")
                            logger.info(f"  【3】目标达成情况:")
                            logger.info(f"      - 目标 AFP: {target_afp_from_excel:.2f}")
                            logger.info(f"      - 扩展后 AFP: {new_afp_total:.2f}")
                            logger.info(f"      - 剩余差距：{target_afp_from_excel - new_afp_total:.2f}")
                                                    
                            if new_afp_total >= target_afp_from_excel:
                                logger.info(f"      ✅ 成功！扩展后 AFP ({new_afp_total:.2f}) >= 目标 AFP ({target_afp_from_excel:.2f})")
                                logger.info(f"      🎉 超额完成：超出 {new_afp_total - target_afp_from_excel:.2f}")
                            else:
                                remaining_gap = target_afp_from_excel - new_afp_total
                                logger.info(f"      ⚠️ AFP 仍有差距：{remaining_gap:.2f}")
                                logger.info(f"      💡 建议：可能需要手动添加更多功能点或调整现有功能点类别")
                                                    
                            logger.info(f"")
                            logger.info(f"  【4】新增功能点详情 (前 5 个):")
                            for i, point in enumerate(expanded_points[:5], 1):
                                logger.info(f"      [{i}] {point.get('功能点计数项', '')}")
                                logger.info(f"          备注：{point.get('备注', '')}")
                                logger.info(f"          类别：{point.get('类别', '')}")
                                logger.info(f"          UFP: {point.get('UFP', 0)}, AFP: {point.get('AFP', 0):.2f}")
                                                    
                            if len(expanded_points) > 5:
                                logger.info(f"      ... 还有 {len(expanded_points) - 5} 个功能点（详见日志级别 DEBUG）")
                                                    
                            logger.info(f"")
                            logger.info(f"  【5】验证公式:")
                            logger.info(f"      - 新 AFP = 原 AFP + 新增 AFP")
                            logger.info(f"      - {new_afp_total:.2f} = {current_afp:.2f} + {afp_increase:.2f} ✓")
                            logger.info(f"      - 达成率 = 新 AFP / 目标 AFP × 100%")
                            achievement_rate = (new_afp_total / target_afp_from_excel) * 100
                            logger.info(f"      - 达成率 = {new_afp_total:.2f} / {target_afp_from_excel:.2f} × 100% = {achievement_rate:.2f}%")
                            logger.info(f"")
                            logger.info("=" * 80)
                        else:
                            logger.warning("⚠️ [AI_EXPAND] AI 返回空结果，未扩展任何功能点")
                                                            
                    except Exception as ai_error:
                        logger.error(f"❌ [AI_EXPAND] AI 辅助扩展失败：{ai_error}", exc_info=True)
                        logger.warning(f"将使用原始功能点，不进行 AI 扩展")
                else:
                    logger.info(f"❌ 结论：不需要 AI 扩展（当前 AFP {current_afp:.2f} >= 目标 AFP {target_afp_from_excel:.2f}）")
                    logger.info(f"   富余 AFP: {current_afp - target_afp_from_excel:.2f}")
                    logger.info(f"   💡 说明：当前功能点已足够，无需扩展")
                    logger.info("=" * 80)
        except Exception as e:
            logger.warning(f"读取评估结果失败，将使用默认 UFP 值：{e}")
        
        # 生成 Excel 文件（文件名使用可读的时间戳格式：YYYYMMDD_HHMMSS）
        timestamp_str = datetime.fromtimestamp(timestamp).strftime('%Y%m%d_%H%M%S')
        output_filename = f"{filename}_FPA_Estimation_{timestamp_str}.xlsx"
        output_path = output_dir / output_filename
        generate_fpa_excel(function_points, str(output_path))
        
        # 清理临时文件
        try:
            os.remove(temp_md_path)
        except Exception as e:
            logger.warning(f"临时文件删除失败：{e}")
        
        flash(f"FPA预估表生成成功！共提取 {len(function_points)} 个功能点", "success")
        
        # 提供下载链接
        return render_template(
            'fpa_generator.html',
            success=True,
            output_file=output_filename,
            output_path=str(output_path),
            function_point_count=len(function_points)
        )
        
    except Exception as e:
        logger.error(f"FPA 生成失败：{e}", exc_info=True)
        flash(f"生成失败：{str(e)}", "error")
        return render_template('fpa_generator.html')


@fpa_generator_bp.route('/download/<filename>')
def download_fpa(filename):
    """下载生成的 FPA预估表"""
    try:
        output_dir = Path(current_app.config['UPLOAD_FOLDER']) / "fpa_output"
        file_path = output_dir / filename
        
        if not file_path.exists():
            flash("文件不存在或已过期！", "error")
            return render_template('fpa_generator.html')
        
        return send_file(
            str(file_path),
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"文件下载失败：{e}")
        flash("下载失败！", "error")
        return render_template('fpa_generator.html')


def get_db_connection():
    """获取数据库连接"""
    return mysql.connector.connect(
        host=current_app.config['MYSQL_HOST'],
        port=current_app.config['MYSQL_PORT'],
        user=current_app.config['MYSQL_USER'],
        password=current_app.config['MYSQL_PASSWORD'],
        charset=current_app.config['MYSQL_CHARSET'],
        database=current_app.config['KNOWLEDGE_BASE_DB']
    )


@fpa_generator_bp.route('/api/evaluation-result/calculate', methods=['POST'])
def calculate_evaluation_result_from_effort():
    """根据调整后工作量反向计算评估结果"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'message': '未提供数据'
            }), 400
        
        # 获取输入参数
        adjusted_effort = float(data.get('adjusted_effort', 0))
        
        # 获取其他固定参数
        cf = float(data.get('cf', WORK_PARAMS['规模变更调整因子']))
        base_productivity = float(data.get('base_productivity', WORK_PARAMS['基准生产率']))
        factor_a = float(data.get('factor_application_type', WORK_PARAMS['调整因子']['应用类型']))
        factor_b = float(data.get('factor_quality', WORK_PARAMS['调整因子']['质量特性']))
        factor_c = float(data.get('factor_language', WORK_PARAMS['调整因子']['开发语言']))
        factor_d = float(data.get('factor_team', WORK_PARAMS['调整因子']['开发团队背景']))
        
        # 反向计算公式：
        # 调整后工作量 = 未调整工作量 × A × B × C × D
        # 未调整工作量 = 调整后规模 × PDR ÷ 8
        # 调整后规模 = AFP × CF
        # AFP ≈ UFP（简化计算）
        
        total_factor = factor_a * factor_b * factor_c * factor_d
        
        # 1. 计算未调整工作量
        unadjusted_effort = adjusted_effort / total_factor
        
        # 2. 计算调整后规模
        adjusted_scale = (unadjusted_effort * 8) / base_productivity
        
        # 3. 计算 AFP
        afp = adjusted_scale / cf
        
        # 4. 计算 UFP（AFP ≈ UFP，简化计算）
        ufp = afp
        
        logger.info(f"反向计算：调整后工作量={adjusted_effort}, UFP={ufp:.2f}")
        
        result = {
            'afp': round(afp, 2),
            'cf': cf,
            'adjusted_scale': round(adjusted_scale, 2),
            'base_productivity': base_productivity,
            'unadjusted_effort': round(unadjusted_effort, 2),
            'factor_application_type': factor_a,
            'factor_quality': factor_b,
            'factor_language': factor_c,
            'factor_team': factor_d,
            'adjusted_effort': round(adjusted_effort, 2),
            'total_factor': round(total_factor, 2),
            'target_ufp': round(ufp, 2)
        }
        
        return jsonify({
            'success': True,
            'data': result
        })
        
    except Exception as e:
        logger.error(f"反向计算失败：{e}", exc_info=True)
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@fpa_generator_bp.route('/api/evaluation-result', methods=['GET'])
def get_evaluation_result():
    """获取评估结果数据"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # 查询最新的评估结果
        cursor.execute("""
            SELECT * FROM fpa_evaluation_result
            ORDER BY created_at DESC
            LIMIT 1
        """)
        
        result = cursor.fetchone()
        
        # 转换 Decimal 为 float
        if result:
            for key, value in result.items():
                if isinstance(value, Decimal):
                    result[key] = float(value)
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'data': result
        })
        
    except Exception as e:
        logger.error(f"获取评估结果失败：{e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@fpa_generator_bp.route('/api/evaluation-result/from-excel', methods=['GET'])
def get_evaluation_result_from_excel():
    """从最新生成的 Excel 文件中读取评估结果"""
    try:
        # 获取最新的 FPA 输出文件
        output_dir = Path(current_app.config['UPLOAD_FOLDER']) / "fpa_output"
        excel_files = list(output_dir.glob("*.xlsx"))
        
        if not excel_files:
            return jsonify({
                'success': False,
                'message': '未找到 Excel 文件'
            }), 404
        
        # 按修改时间排序，获取最新文件
        latest_file = max(excel_files, key=lambda f: f.stat().st_mtime)
        logger.info(f"读取最新 Excel 文件：{latest_file}")
        
        # 读取"2. 规模估算"工作表
        df_scale = pd.read_excel(latest_file, sheet_name="2. 规模估算")
        
        # 读取"3. 调整因子"工作表
        df_adjustment = pd.read_excel(latest_file, sheet_name="3. 调整因子")
        
        # 从规模估算表获取 B4 单元格的值（调整后规模）
        # 假设 B 列是数值列
        if len(df_scale) >= 4:
            # 获取第 4 行（索引 3）的 B 列值
            afp_value = df_scale.iloc[3, 1]  # B4 单元格
        else:
            # 如果行数不足，尝试计算 AFP 总和
            if 'AFP' in df_scale.columns:
                afp_value = df_scale['AFP'].sum()
            else:
                afp_value = 0
        
        # 从调整因子表获取各项调整因子
        adjustment_factors = {
            '应用类型': 1.00,
            '质量特性': 0.90,
            '开发语言': 1.00,
            '开发团队背景': 0.80
        }
        
        # 尝试从调整因子表中读取
        if len(df_adjustment) > 0:
            for _, row in df_adjustment.iterrows():
                if '调整因子名称' in df_adjustment.columns and '值' in df_adjustment.columns:
                    factor_name = row['调整因子名称']
                    factor_value = row['值']
                    if factor_name in adjustment_factors:
                        adjustment_factors[factor_name] = float(factor_value)
        
        # 获取规模变更调整因子 CF
        cf_value = WORK_PARAMS.get('规模变更调整因子', 1.21)
        
        # 获取基准生产率 PDR
        pdr_value = WORK_PARAMS.get('基准生产率', 10.12)
        
        # 计算各项值
        afp = float(afp_value) if pd.notna(afp_value) else 0
        cf = float(cf_value)
        adjusted_scale = afp * cf  # 调整后规模
        base_productivity = float(pdr_value)
        unadjusted_effort = (adjusted_scale * base_productivity) / 8  # 未调整工作量
        
        # 计算调整因子乘积
        total_factor = 1.0
        for factor_value in adjustment_factors.values():
            total_factor *= float(factor_value)
        
        adjusted_effort = unadjusted_effort * total_factor  # 调整后工作量
        
        # 计算用例数量（根据 AFP 值计算）
        # 规则：AFP 值越大，生成的用例数量越多
        # 建议：每 2-3 个功能点生成 1 个用例
        test_case_count = max(1, round(afp / 2.5))
        
        evaluation_data = {
            'afp': afp,
            'cf': cf,
            'adjusted_scale': adjusted_scale,
            'base_productivity': base_productivity,
            'unadjusted_effort': unadjusted_effort,
            'factor_application_type': adjustment_factors['应用类型'],
            'factor_quality': adjustment_factors['质量特性'],
            'factor_language': adjustment_factors['开发语言'],
            'factor_team': adjustment_factors['开发团队背景'],
            'adjusted_effort': adjusted_effort,
            'total_factor': total_factor,
            'test_case_count': test_case_count
        }
        
        logger.info(f"从 Excel 读取评估结果：{evaluation_data}")
        
        return jsonify({
            'success': True,
            'data': evaluation_data,
            'file': str(latest_file.name)
        })
        
    except Exception as e:
        logger.error(f"从 Excel 读取评估结果失败：{e}", exc_info=True)
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@fpa_generator_bp.route('/api/evaluation-result', methods=['POST'])
def save_evaluation_result():
    """保存评估结果数据"""
    try:
        data = request.get_json()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 检查是否已存在
        cursor.execute("""
            SELECT id FROM fpa_evaluation_result
            WHERE requirement_code = %s
        """, (data.get('requirement_code'),))
        
        existing = cursor.fetchone()
        
        if existing:
            # 更新
            cursor.execute("""
                UPDATE fpa_evaluation_result
                SET 
                    requirement_name = %s,
                    afp = %s,
                    cf = %s,
                    adjusted_scale = %s,
                    base_productivity = %s,
                    unadjusted_effort = %s,
                    factor_application_type = %s,
                    factor_quality = %s,
                    factor_language = %s,
                    factor_team = %s,
                    adjusted_effort = %s,
                    updated_at = NOW()
                WHERE requirement_code = %s
            """, (
                data.get('requirement_name'),
                data.get('afp'),
                data.get('cf'),
                data.get('adjusted_scale'),
                data.get('base_productivity'),
                data.get('unadjusted_effort'),
                data.get('factor_application_type'),
                data.get('factor_quality'),
                data.get('factor_language'),
                data.get('factor_team'),
                data.get('adjusted_effort'),
                data.get('requirement_code')
            ))
            logger.info(f"更新评估结果：{data.get('requirement_code')}")
        else:
            # 插入
            cursor.execute("""
                INSERT INTO fpa_evaluation_result (
                    requirement_name, requirement_code,
                    afp, cf, adjusted_scale, base_productivity, unadjusted_effort,
                    factor_application_type, factor_quality, factor_language, factor_team,
                    adjusted_effort
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                data.get('requirement_name'),
                data.get('requirement_code'),
                data.get('afp'),
                data.get('cf'),
                data.get('adjusted_scale'),
                data.get('base_productivity'),
                data.get('unadjusted_effort'),
                data.get('factor_application_type'),
                data.get('factor_quality'),
                data.get('factor_language'),
                data.get('factor_team'),
                data.get('adjusted_effort')
            ))
            logger.info(f"插入评估结果：{data.get('requirement_code')}")
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '评估结果已保存'
        })
        
    except Exception as e:
        logger.error(f"保存评估结果失败：{e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500



